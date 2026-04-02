"""
Module Comptabilité — Journal, Grand Livre, Balance, CPC, Bilan
Lecture seule — données calculées depuis Depense, Recette, Paiement, CaisseMouvement
"""
import io
import datetime
import re
from decimal import Decimal

from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from django.http import HttpResponse
from django.db.models import Sum

from .models import Depense, Recette, Paiement, CompteComptable, DetailAppelCharge, AppelCharge
from .views import get_user_residence


# ============================================================
# Comptes virtuels
# ============================================================
COMPTE_TRESORERIE = ("512", "Banque / Trésorerie")
COMPTE_COPRO      = ("342", "Copropriétaires")


# ============================================================
# Helper principal — construction des écritures du journal
# ============================================================
def _build_entries(residence, date_debut=None, date_fin=None):
    """
    Retourne une liste d'entrées de journal (double-entrée simplifiée).
    Chaque entrée : {date, type, piece, libelle, compte_code, compte_libelle, debit, credit}
    """
    entries = []
    D0 = Decimal("0")

    # ── Dépenses ────────────────────────────────────────────
    dep_qs = (
        Depense.objects
        .select_related("compte", "fournisseur")
        .filter(residence=residence)
    )
    if date_debut:
        dep_qs = dep_qs.filter(date_depense__gte=date_debut)
    if date_fin:
        dep_qs = dep_qs.filter(date_depense__lte=date_fin)

    for d in dep_qs:
        date  = d.date_depense
        piece = d.facture_reference or ""
        lib   = d.libelle or ""
        m     = d.montant or D0
        c_code = d.compte.code    if d.compte else "6XX"
        c_lib  = d.compte.libelle if d.compte else "Compte charge"
        # Débit compte charge
        entries.append(dict(date=date, type="Dépense", piece=piece, libelle=lib,
                            compte_code=c_code, compte_libelle=c_lib,
                            debit=m, credit=D0))
        # Crédit trésorerie
        entries.append(dict(date=date, type="Dépense", piece=piece, libelle=lib,
                            compte_code=COMPTE_TRESORERIE[0],
                            compte_libelle=COMPTE_TRESORERIE[1],
                            debit=D0, credit=m))

    # ── Recettes ────────────────────────────────────────────
    rec_qs = (
        Recette.objects
        .select_related("compte")
        .filter(residence=residence)
    )
    if date_debut:
        rec_qs = rec_qs.filter(date_recette__gte=date_debut)
    if date_fin:
        rec_qs = rec_qs.filter(date_recette__lte=date_fin)

    for r in rec_qs:
        date  = r.date_recette
        piece = ""
        lib   = r.libelle or ""
        m     = r.montant or D0
        c_code = r.compte.code    if r.compte else "7XX"
        c_lib  = r.compte.libelle if r.compte else "Compte produit"
        # Débit trésorerie
        entries.append(dict(date=date, type="Recette", piece=piece, libelle=lib,
                            compte_code=COMPTE_TRESORERIE[0],
                            compte_libelle=COMPTE_TRESORERIE[1],
                            debit=m, credit=D0))
        # Crédit compte produit
        entries.append(dict(date=date, type="Recette", piece=piece, libelle=lib,
                            compte_code=c_code, compte_libelle=c_lib,
                            debit=D0, credit=m))

    # ── Paiements ───────────────────────────────────────────
    pai_qs = (
        Paiement.objects
        .select_related("lot__representant")
        .filter(residence=residence)
    )
    if date_debut:
        pai_qs = pai_qs.filter(date_paiement__gte=date_debut)
    if date_fin:
        pai_qs = pai_qs.filter(date_paiement__lte=date_fin)

    for p in pai_qs:
        date  = p.date_paiement
        piece = p.reference or ""
        rep   = p.lot.representant if p.lot and p.lot.representant else None
        lot_n = p.lot.numero_lot if p.lot else "?"
        lib   = f"Paiement lot {lot_n}"
        if rep:
            lib += f" — {rep.nom} {rep.prenom or ''}".rstrip()
        m = p.montant or D0
        # Débit trésorerie
        entries.append(dict(date=date, type="Paiement", piece=piece, libelle=lib,
                            compte_code=COMPTE_TRESORERIE[0],
                            compte_libelle=COMPTE_TRESORERIE[1],
                            debit=m, credit=D0))
        # Crédit copropriétaires
        entries.append(dict(date=date, type="Paiement", piece=piece, libelle=lib,
                            compte_code=COMPTE_COPRO[0],
                            compte_libelle=COMPTE_COPRO[1],
                            debit=D0, credit=m))

    # Tri chronologique
    entries.sort(key=lambda e: e["date"])
    return entries


def _filter_entries_a_affecter(entries):
    """Keep only entries linked to account 000 (unclassified)."""
    return [e for e in entries if e["compte_code"] == "000"]


def _serialize_entry(e):
    return {
        "date":            str(e["date"]),
        "type":            e["type"],
        "piece":           e["piece"],
        "libelle":         e["libelle"],
        "compte_code":     e["compte_code"],
        "compte_libelle":  e["compte_libelle"],
        "debit":           str(e["debit"]),
        "credit":          str(e["credit"]),
    }


# ============================================================
# Journal
# ============================================================
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def journal_view(request):
    residence = get_user_residence(request)
    if not residence:
        return Response({"detail": "Aucune résidence."}, status=400)

    date_debut   = request.query_params.get("date_debut")   or None
    date_fin     = request.query_params.get("date_fin")     or None
    a_affecter   = request.query_params.get("a_affecter")  == "true"
    entries      = _build_entries(residence, date_debut, date_fin)

    if a_affecter:
        entries = _filter_entries_a_affecter(entries)

    total_debit  = sum(e["debit"]  for e in entries)
    total_credit = sum(e["credit"] for e in entries)

    return Response({
        "entries":       [_serialize_entry(e) for e in entries],
        "total_debit":   str(total_debit),
        "total_credit":  str(total_credit),
        "a_affecter":    a_affecter,
    })


# ============================================================
# Grand Livre
# ============================================================
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def grand_livre_view(request):
    residence   = get_user_residence(request)
    if not residence:
        return Response({"detail": "Aucune résidence."}, status=400)

    date_debut    = request.query_params.get("date_debut")    or None
    date_fin      = request.query_params.get("date_fin")      or None
    filter_compte = request.query_params.get("compte")        or None
    entries       = _build_entries(residence, date_debut, date_fin)

    # Group by compte_code
    from collections import OrderedDict
    comptes = OrderedDict()
    for e in entries:
        key = (e["compte_code"], e["compte_libelle"])
        if key not in comptes:
            comptes[key] = []
        comptes[key].append(e)

    result = []
    for (code, libelle), lines in sorted(comptes.items()):
        if filter_compte and filter_compte.lower() not in code.lower() and filter_compte.lower() not in libelle.lower():
            continue
        solde = Decimal("0")
        compte_entries = []
        total_d = Decimal("0")
        total_c = Decimal("0")
        for e in lines:
            solde += e["debit"] - e["credit"]
            total_d += e["debit"]
            total_c += e["credit"]
            compte_entries.append({
                **_serialize_entry(e),
                "solde": str(solde),
            })
        result.append({
            "compte_code":    code,
            "compte_libelle": libelle,
            "total_debit":    str(total_d),
            "total_credit":   str(total_c),
            "solde":          str(total_d - total_c),
            "entries":        compte_entries,
        })

    return Response({"comptes": result})


# ============================================================
# Balance
# ============================================================
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def balance_view(request):
    residence  = get_user_residence(request)
    if not residence:
        return Response({"detail": "Aucune résidence."}, status=400)

    date_debut = request.query_params.get("date_debut") or None
    date_fin   = request.query_params.get("date_fin")   or None
    entries    = _build_entries(residence, date_debut, date_fin)

    from collections import defaultdict
    totals = defaultdict(lambda: {"libelle": "", "debit": Decimal("0"), "credit": Decimal("0")})
    for e in entries:
        code = e["compte_code"]
        totals[code]["libelle"] = e["compte_libelle"]
        totals[code]["debit"]  += e["debit"]
        totals[code]["credit"] += e["credit"]

    rows = []
    grand_d = Decimal("0")
    grand_c = Decimal("0")
    for code in sorted(totals.keys()):
        t = totals[code]
        d, c = t["debit"], t["credit"]
        grand_d += d
        grand_c += c
        rows.append({
            "compte_code":    code,
            "compte_libelle": t["libelle"],
            "total_debit":    str(d),
            "total_credit":   str(c),
            "solde_debiteur":  str(max(d - c, Decimal("0"))),
            "solde_crediteur": str(max(c - d, Decimal("0"))),
        })

    return Response({
        "rows":          rows,
        "grand_debit":   str(grand_d),
        "grand_credit":  str(grand_c),
    })


# ============================================================
# CPC — Compte de Produits et Charges
# ============================================================
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def cpc_view(request):
    residence  = get_user_residence(request)
    if not residence:
        return Response({"detail": "Aucune résidence."}, status=400)

    date_debut = request.query_params.get("date_debut") or None
    date_fin   = request.query_params.get("date_fin")   or None
    D0 = Decimal("0")

    # ── Charges (dépenses par compte) ───────────────────────
    dep_qs = Depense.objects.select_related("compte").filter(residence=residence)
    if date_debut:
        dep_qs = dep_qs.filter(date_depense__gte=date_debut)
    if date_fin:
        dep_qs = dep_qs.filter(date_depense__lte=date_fin)

    charges_map = {}
    for d in dep_qs:
        code = d.compte.code    if d.compte else "6XX"
        lib  = d.compte.libelle if d.compte else "Divers charges"
        key  = (code, lib)
        charges_map[key] = charges_map.get(key, D0) + (d.montant or D0)

    charges = sorted(
        [{"compte_code": k[0], "compte_libelle": k[1], "montant": str(v)}
         for k, v in charges_map.items()],
        key=lambda x: x["compte_code"]
    )
    total_charges = sum(charges_map.values())

    # ── Produits (recettes par compte) ──────────────────────
    rec_qs = Recette.objects.select_related("compte").filter(residence=residence)
    if date_debut:
        rec_qs = rec_qs.filter(date_recette__gte=date_debut)
    if date_fin:
        rec_qs = rec_qs.filter(date_recette__lte=date_fin)

    produits_map = {}
    for r in rec_qs:
        code = r.compte.code    if r.compte else "7XX"
        lib  = r.compte.libelle if r.compte else "Divers produits"
        key  = (code, lib)
        produits_map[key] = produits_map.get(key, D0) + (r.montant or D0)

    # Ajouter les paiements copropriétaires comme produit (342)
    pai_qs = Paiement.objects.filter(residence=residence)
    if date_debut:
        pai_qs = pai_qs.filter(date_paiement__gte=date_debut)
    if date_fin:
        pai_qs = pai_qs.filter(date_paiement__lte=date_fin)
    total_paiements = pai_qs.aggregate(t=Sum("montant"))["t"] or D0
    if total_paiements:
        key342 = (COMPTE_COPRO[0], "Charges récupérées — copropriétaires")
        produits_map[key342] = produits_map.get(key342, D0) + total_paiements

    produits = sorted(
        [{"compte_code": k[0], "compte_libelle": k[1], "montant": str(v)}
         for k, v in produits_map.items()],
        key=lambda x: x["compte_code"]
    )
    total_produits = sum(produits_map.values())

    resultat = total_produits - total_charges

    return Response({
        "charges":        charges,
        "total_charges":  str(total_charges),
        "produits":       produits,
        "total_produits": str(total_produits),
        "resultat":       str(resultat),
    })


# ============================================================
# Bilan simplifié
# ============================================================
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def bilan_view(request):
    residence  = get_user_residence(request)
    if not residence:
        return Response({"detail": "Aucune résidence."}, status=400)

    date_debut = request.query_params.get("date_debut") or None
    date_fin   = request.query_params.get("date_fin")   or None
    D0 = Decimal("0")

    # ── ACTIF ───────────────────────────────────────────────

    # 1. Trésorerie nette (Recettes + Paiements - Dépenses)
    dep_qs = Depense.objects.filter(residence=residence)
    rec_qs = Recette.objects.filter(residence=residence)
    pai_qs = Paiement.objects.filter(residence=residence)
    if date_debut:
        dep_qs = dep_qs.filter(date_depense__gte=date_debut)
        rec_qs = rec_qs.filter(date_recette__gte=date_debut)
        pai_qs = pai_qs.filter(date_paiement__gte=date_debut)
    if date_fin:
        dep_qs = dep_qs.filter(date_depense__lte=date_fin)
        rec_qs = rec_qs.filter(date_recette__lte=date_fin)
        pai_qs = pai_qs.filter(date_paiement__lte=date_fin)

    total_depenses  = dep_qs.aggregate(t=Sum("montant"))["t"] or D0
    total_recettes  = rec_qs.aggregate(t=Sum("montant"))["t"] or D0
    total_paiements = pai_qs.aggregate(t=Sum("montant"))["t"] or D0
    tresorerie_nette = total_recettes + total_paiements - total_depenses

    # 2. Créances copropriétaires (appels non encore payés)
    #    = sum détails_appel.montant - montant_recu where statut != PAYE
    detail_qs = DetailAppelCharge.objects.filter(
        appel__residence=residence
    ).exclude(statut="PAYE")
    if date_debut:
        detail_qs = detail_qs.filter(appel__date_emission__gte=date_debut)
    if date_fin:
        detail_qs = detail_qs.filter(appel__date_emission__lte=date_fin)

    agg = detail_qs.aggregate(
        du=Sum("montant"),
        recu=Sum("montant_recu"),
    )
    creances = (agg["du"] or D0) - (agg["recu"] or D0)
    if creances < D0:
        creances = D0

    total_actif = tresorerie_nette + creances

    # ── PASSIF ──────────────────────────────────────────────

    # 1. Résultat net (Recettes externes - Dépenses)
    resultat_net = total_recettes - total_depenses

    # 2. Fonds travaux collectés (AppelCharge type FOND — montants reçus)
    fond_qs = DetailAppelCharge.objects.filter(
        appel__residence=residence,
        appel__type_charge="FOND",
    )
    if date_debut:
        fond_qs = fond_qs.filter(appel__date_emission__gte=date_debut)
    if date_fin:
        fond_qs = fond_qs.filter(appel__date_emission__lte=date_fin)
    fonds_travaux = fond_qs.aggregate(t=Sum("montant_recu"))["t"] or D0

    # 3. Avances copropriétaires (total paiements reçus)
    avances_copro = total_paiements

    total_passif = resultat_net + fonds_travaux + avances_copro

    return Response({
        "actif": {
            "tresorerie_nette": str(tresorerie_nette),
            "creances_copro":   str(creances),
            "total":            str(total_actif),
        },
        "passif": {
            "resultat_net":     str(resultat_net),
            "fonds_travaux":    str(fonds_travaux),
            "avances_copro":    str(avances_copro),
            "total":            str(total_passif),
        },
    })


# ============================================================
# Exports Excel
# ============================================================
def _excel_header(ws, headers, widths):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    thin   = Side(style="thin", color="E2E8F0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        ws.column_dimensions[chr(64 + col)].width = w
        c = ws.cell(1, col, h)
        c.fill      = PatternFill("solid", fgColor="0F172A")
        c.font      = Font(bold=True, color="FFFFFF", size=10)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = border

def _excel_row(ws, row, values, fmt=None, bold=False):
    from openpyxl.styles import Font, Border, Side
    thin   = Side(style="thin", color="E2E8F0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col, val in enumerate(values, 1):
        c = ws.cell(row, col, val)
        c.border = border
        if fmt and col >= len(values) - 1:
            c.number_format = fmt
        if bold:
            c.font = Font(bold=True)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def journal_export_excel(request):
    from openpyxl import Workbook
    residence = get_user_residence(request)
    if not residence:
        return Response({"detail": "Aucune résidence."}, status=400)
    date_debut = request.query_params.get("date_debut") or None
    date_fin   = request.query_params.get("date_fin")   or None
    entries    = _build_entries(residence, date_debut, date_fin)

    wb = Workbook()
    ws = wb.active
    ws.title = "Journal"
    _excel_header(ws, ["Date", "Type", "Pièce", "Libellé", "Compte", "Intitulé", "Débit", "Crédit"],
                      [12,     14,     14,       36,        10,       28,         14,      14])
    for i, e in enumerate(entries, 2):
        from openpyxl.styles import Border, Side
        thin   = Side(style="thin", color="E2E8F0")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for col, val in enumerate([
            str(e["date"]), e["type"], e["piece"], e["libelle"],
            e["compte_code"], e["compte_libelle"],
            float(e["debit"]) or None, float(e["credit"]) or None,
        ], 1):
            c = ws.cell(i, col, val)
            c.border = border
            if col in (7, 8) and val is not None:
                c.number_format = "#,##0.00"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    slug  = re.sub(r"[^a-z0-9]", "_", residence.nom_residence.lower())
    today = datetime.date.today().strftime("%Y%m%d")
    resp  = HttpResponse(buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = f'attachment; filename="journal_{slug}_{today}.xlsx"'
    return resp


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def balance_export_excel(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Border, Side
    from collections import defaultdict

    residence = get_user_residence(request)
    if not residence:
        return Response({"detail": "Aucune résidence."}, status=400)
    date_debut = request.query_params.get("date_debut") or None
    date_fin   = request.query_params.get("date_fin")   or None
    entries    = _build_entries(residence, date_debut, date_fin)

    totals = defaultdict(lambda: {"libelle": "", "debit": Decimal("0"), "credit": Decimal("0")})
    for e in entries:
        code = e["compte_code"]
        totals[code]["libelle"] = e["compte_libelle"]
        totals[code]["debit"]  += e["debit"]
        totals[code]["credit"] += e["credit"]

    thin   = Side(style="thin", color="E2E8F0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = Workbook()
    ws = wb.active
    ws.title = "Balance"
    _excel_header(ws, ["Compte", "Intitulé", "Total débit", "Total crédit", "Solde débiteur", "Solde créditeur"],
                      [10,       32,          16,            16,             16,               16])
    grand_d = grand_c = Decimal("0")
    for i, code in enumerate(sorted(totals.keys()), 2):
        t = totals[code]
        d, c = t["debit"], t["credit"]
        grand_d += d; grand_c += c
        sd = float(max(d - c, Decimal("0")))
        sc = float(max(c - d, Decimal("0")))
        for col, val in enumerate([code, t["libelle"], float(d), float(c), sd or None, sc or None], 1):
            cell = ws.cell(i, col, val)
            cell.border = border
            if col >= 3 and val is not None:
                cell.number_format = "#,##0.00"
    # Totaux
    last = len(totals) + 2
    ws.cell(last, 1, "TOTAL").font = Font(bold=True)
    ws.cell(last, 3, float(grand_d)).number_format = "#,##0.00"
    ws.cell(last, 3).font = Font(bold=True)
    ws.cell(last, 4, float(grand_c)).number_format = "#,##0.00"
    ws.cell(last, 4).font = Font(bold=True)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    slug  = re.sub(r"[^a-z0-9]", "_", residence.nom_residence.lower())
    today = datetime.date.today().strftime("%Y%m%d")
    resp  = HttpResponse(buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = f'attachment; filename="balance_{slug}_{today}.xlsx"'
    return resp


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def cpc_export_excel(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side
    from django.db.models import Sum

    residence = get_user_residence(request)
    if not residence:
        return Response({"detail": "Aucune résidence."}, status=400)
    date_debut = request.query_params.get("date_debut") or None
    date_fin   = request.query_params.get("date_fin")   or None
    D0 = Decimal("0")

    dep_qs = Depense.objects.select_related("compte").filter(residence=residence)
    rec_qs = Recette.objects.select_related("compte").filter(residence=residence)
    pai_qs = Paiement.objects.filter(residence=residence)
    if date_debut:
        dep_qs = dep_qs.filter(date_depense__gte=date_debut)
        rec_qs = rec_qs.filter(date_recette__gte=date_debut)
        pai_qs = pai_qs.filter(date_paiement__gte=date_debut)
    if date_fin:
        dep_qs = dep_qs.filter(date_depense__lte=date_fin)
        rec_qs = rec_qs.filter(date_recette__lte=date_fin)
        pai_qs = pai_qs.filter(date_paiement__lte=date_fin)

    charges_map = {}
    for d in dep_qs:
        key = (d.compte.code if d.compte else "6XX", d.compte.libelle if d.compte else "Divers")
        charges_map[key] = charges_map.get(key, D0) + (d.montant or D0)

    produits_map = {}
    for r in rec_qs:
        key = (r.compte.code if r.compte else "7XX", r.compte.libelle if r.compte else "Divers")
        produits_map[key] = produits_map.get(key, D0) + (r.montant or D0)
    total_pai = pai_qs.aggregate(t=Sum("montant"))["t"] or D0
    if total_pai:
        produits_map[(COMPTE_COPRO[0], "Charges récupérées — copropriétaires")] = \
            produits_map.get((COMPTE_COPRO[0], ""), D0) + total_pai

    total_charges  = sum(charges_map.values())
    total_produits = sum(produits_map.values())
    resultat       = total_produits - total_charges

    thin   = Side(style="thin", color="E2E8F0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = Workbook()
    ws = wb.active
    ws.title = "CPC"
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 18

    row = 1
    ws.cell(row, 1, "COMPTE PRODUITS ET CHARGES").font = Font(bold=True, size=13)
    row += 2

    ws.cell(row, 1, "CHARGES").font = Font(bold=True, color="DC2626")
    row += 1
    for (code, lib), m in sorted(charges_map.items()):
        ws.cell(row, 1, code); ws.cell(row, 2, lib)
        c = ws.cell(row, 3, float(m)); c.number_format = "#,##0.00"
        row += 1
    ws.cell(row, 2, "Total charges").font = Font(bold=True)
    c = ws.cell(row, 3, float(total_charges)); c.number_format = "#,##0.00"; c.font = Font(bold=True)
    row += 2

    ws.cell(row, 1, "PRODUITS").font = Font(bold=True, color="059669")
    row += 1
    for (code, lib), m in sorted(produits_map.items()):
        ws.cell(row, 1, code); ws.cell(row, 2, lib)
        c = ws.cell(row, 3, float(m)); c.number_format = "#,##0.00"
        row += 1
    ws.cell(row, 2, "Total produits").font = Font(bold=True)
    c = ws.cell(row, 3, float(total_produits)); c.number_format = "#,##0.00"; c.font = Font(bold=True)
    row += 2

    ws.cell(row, 2, "RÉSULTAT NET").font = Font(bold=True)
    c = ws.cell(row, 3, float(resultat)); c.number_format = "#,##0.00"
    c.font = Font(bold=True, color="059669" if resultat >= 0 else "DC2626")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    slug  = re.sub(r"[^a-z0-9]", "_", residence.nom_residence.lower())
    today = datetime.date.today().strftime("%Y%m%d")
    resp  = HttpResponse(buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = f'attachment; filename="cpc_{slug}_{today}.xlsx"'
    return resp


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def bilan_export_excel(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from django.db.models import Sum

    residence = get_user_residence(request)
    if not residence:
        return Response({"detail": "Aucune résidence."}, status=400)
    date_debut = request.query_params.get("date_debut") or None
    date_fin   = request.query_params.get("date_fin")   or None
    D0 = Decimal("0")

    dep_qs = Depense.objects.filter(residence=residence)
    rec_qs = Recette.objects.filter(residence=residence)
    pai_qs = Paiement.objects.filter(residence=residence)
    if date_debut:
        dep_qs = dep_qs.filter(date_depense__gte=date_debut)
        rec_qs = rec_qs.filter(date_recette__gte=date_debut)
        pai_qs = pai_qs.filter(date_paiement__gte=date_debut)
    if date_fin:
        dep_qs = dep_qs.filter(date_depense__lte=date_fin)
        rec_qs = rec_qs.filter(date_recette__lte=date_fin)
        pai_qs = pai_qs.filter(date_paiement__lte=date_fin)

    total_dep  = dep_qs.aggregate(t=Sum("montant"))["t"] or D0
    total_rec  = rec_qs.aggregate(t=Sum("montant"))["t"] or D0
    total_pai  = pai_qs.aggregate(t=Sum("montant"))["t"] or D0
    tresorerie = total_rec + total_pai - total_dep

    detail_qs = DetailAppelCharge.objects.filter(appel__residence=residence).exclude(statut="PAYE")
    if date_debut:
        detail_qs = detail_qs.filter(appel__date_emission__gte=date_debut)
    if date_fin:
        detail_qs = detail_qs.filter(appel__date_emission__lte=date_fin)
    agg = detail_qs.aggregate(du=Sum("montant"), recu=Sum("montant_recu"))
    creances = max((agg["du"] or D0) - (agg["recu"] or D0), D0)

    fond_qs = DetailAppelCharge.objects.filter(appel__residence=residence, appel__type_charge="FOND")
    if date_debut:
        fond_qs = fond_qs.filter(appel__date_emission__gte=date_debut)
    if date_fin:
        fond_qs = fond_qs.filter(appel__date_emission__lte=date_fin)
    fonds = fond_qs.aggregate(t=Sum("montant_recu"))["t"] or D0

    resultat = total_rec - total_dep

    wb = Workbook()
    ws = wb.active
    ws.title = "Bilan"
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["D"].width = 36
    ws.column_dimensions["E"].width = 20

    row = 1
    ws.cell(row, 1, "BILAN SIMPLIFIÉ").font = Font(bold=True, size=13)
    row += 2

    ws.cell(row, 1, "ACTIF").font = Font(bold=True, color="1d4ed8")
    ws.cell(row, 4, "PASSIF").font = Font(bold=True, color="059669")
    row += 1

    actif_rows = [
        ("Trésorerie nette (512)", tresorerie),
        ("Créances copropriétaires (342)", creances),
        ("TOTAL ACTIF", tresorerie + creances),
    ]
    passif_rows = [
        ("Résultat net de l'exercice", resultat),
        ("Fonds travaux collectés", fonds),
        ("Avances copropriétaires", total_pai),
        ("TOTAL PASSIF", resultat + fonds + total_pai),
    ]

    for i, ((al, av), (pl, pv)) in enumerate(
        zip(actif_rows + [(None, None)] * (len(passif_rows) - len(actif_rows)),
            passif_rows)
    ):
        if al:
            bold = al.startswith("TOTAL")
            ws.cell(row + i, 1, al).font = Font(bold=bold)
            c = ws.cell(row + i, 2, float(av)); c.number_format = "#,##0.00"
            if bold: c.font = Font(bold=True)
        if pl:
            bold = pl.startswith("TOTAL")
            ws.cell(row + i, 4, pl).font = Font(bold=bold)
            c = ws.cell(row + i, 5, float(pv)); c.number_format = "#,##0.00"
            if bold: c.font = Font(bold=True)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    slug  = re.sub(r"[^a-z0-9]", "_", residence.nom_residence.lower())
    today = datetime.date.today().strftime("%Y%m%d")
    resp  = HttpResponse(buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = f'attachment; filename="bilan_{slug}_{today}.xlsx"'
    return resp


# ============================================================
# Exports PDF (Journal & Balance)
# ============================================================
def _pdf_base(buf, residence, title, subtitle=""):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=1.5*cm, rightMargin=1.5*cm,
        topMargin=1.8*cm, bottomMargin=1.8*cm,
    )
    return doc


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def journal_export_pdf(request):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER

    residence = get_user_residence(request)
    if not residence:
        return Response({"detail": "Aucune résidence."}, status=400)
    date_debut = request.query_params.get("date_debut") or None
    date_fin   = request.query_params.get("date_fin")   or None
    entries    = _build_entries(residence, date_debut, date_fin)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
        leftMargin=1.5*cm, rightMargin=1.5*cm, topMargin=1.8*cm, bottomMargin=1.8*cm)

    styles = getSampleStyleSheet()
    title_s = ParagraphStyle("T", parent=styles["Normal"], fontSize=14, fontName="Helvetica-Bold", spaceAfter=4)
    sub_s   = ParagraphStyle("S", parent=styles["Normal"], fontSize=9, textColor=colors.HexColor("#64748B"), spaceAfter=10)

    DARK  = colors.HexColor("#0F172A")
    LIGHT = colors.HexColor("#F8FAFC")

    ts = TableStyle([
        ("BACKGROUND",   (0,0), (-1,0), DARK),
        ("TEXTCOLOR",    (0,0), (-1,0), colors.white),
        ("FONTNAME",     (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE",     (0,0), (-1,-1), 8),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [LIGHT, colors.white]),
        ("GRID",         (0,0), (-1,-1), 0.4, colors.HexColor("#E2E8F0")),
        ("ALIGN",        (6,0), (7,-1), "RIGHT"),
        ("LEFTPADDING",  (0,0), (-1,-1), 5),
        ("RIGHTPADDING", (0,0), (-1,-1), 5),
        ("TOPPADDING",   (0,0), (-1,-1), 4),
        ("BOTTOMPADDING",(0,0), (-1,-1), 4),
    ])

    def money(v):
        fv = float(v)
        return f"{fv:,.2f}" if fv else ""

    periode = f"{date_debut or '—'} → {date_fin or '—'}"

    rows = [["Date", "Type", "Pièce", "Libellé", "Compte", "Intitulé", "Débit", "Crédit"]]
    for e in entries[:200]:
        rows.append([
            str(e["date"]), e["type"], (e["piece"] or "")[:12],
            (e["libelle"] or "")[:40],
            e["compte_code"], (e["compte_libelle"] or "")[:22],
            money(e["debit"]), money(e["credit"]),
        ])

    elems = [
        Paragraph("Journal comptable", title_s),
        Paragraph(f"{residence.nom_residence} · Période : {periode}", sub_s),
        Table(rows, colWidths=[2.2*cm, 2*cm, 2.2*cm, 6.5*cm, 1.6*cm, 4.5*cm, 2.5*cm, 2.5*cm]),
    ]
    elems[-1].setStyle(ts)
    if len(entries) > 200:
        elems.append(Paragraph(f"… {len(entries)-200} ligne(s) dans l'export Excel.", sub_s))

    doc.build(elems)
    buf.seek(0)
    slug  = re.sub(r"[^a-z0-9]", "_", residence.nom_residence.lower())
    today = datetime.date.today().strftime("%Y%m%d")
    resp  = HttpResponse(buf.read(), content_type="application/pdf")
    resp["Content-Disposition"] = f'attachment; filename="journal_{slug}_{today}.pdf"'
    return resp


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def balance_export_pdf(request):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from collections import defaultdict

    residence = get_user_residence(request)
    if not residence:
        return Response({"detail": "Aucune résidence."}, status=400)
    date_debut = request.query_params.get("date_debut") or None
    date_fin   = request.query_params.get("date_fin")   or None
    entries    = _build_entries(residence, date_debut, date_fin)

    totals = defaultdict(lambda: {"libelle": "", "debit": Decimal("0"), "credit": Decimal("0")})
    for e in entries:
        code = e["compte_code"]
        totals[code]["libelle"] = e["compte_libelle"]
        totals[code]["debit"]  += e["debit"]
        totals[code]["credit"] += e["credit"]

    DARK  = colors.HexColor("#0F172A")
    LIGHT = colors.HexColor("#F8FAFC")
    def money(v): return f"{float(v):,.2f}"

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
        leftMargin=1.5*cm, rightMargin=1.5*cm, topMargin=1.8*cm, bottomMargin=1.8*cm)
    styles = getSampleStyleSheet()
    title_s = ParagraphStyle("T", parent=styles["Normal"], fontSize=14, fontName="Helvetica-Bold", spaceAfter=4)
    sub_s   = ParagraphStyle("S", parent=styles["Normal"], fontSize=9, textColor=colors.HexColor("#64748B"), spaceAfter=10)

    ts = TableStyle([
        ("BACKGROUND",    (0,0), (-1,0), DARK),
        ("TEXTCOLOR",     (0,0), (-1,0), colors.white),
        ("FONTNAME",      (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE",      (0,0), (-1,-1), 9),
        ("ROWBACKGROUNDS",(0,1), (-1,-1), [LIGHT, colors.white]),
        ("GRID",          (0,0), (-1,-1), 0.4, colors.HexColor("#E2E8F0")),
        ("ALIGN",         (2,0), (-1,-1), "RIGHT"),
        ("LEFTPADDING",   (0,0), (-1,-1), 6),
        ("RIGHTPADDING",  (0,0), (-1,-1), 6),
        ("TOPPADDING",    (0,0), (-1,-1), 4),
        ("BOTTOMPADDING", (0,0), (-1,-1), 4),
    ])

    rows = [["Compte", "Intitulé", "Total débit", "Total crédit", "Solde débiteur", "Solde créditeur"]]
    grand_d = grand_c = Decimal("0")
    for code in sorted(totals.keys()):
        t = totals[code]
        d, c = t["debit"], t["credit"]
        grand_d += d; grand_c += c
        sd = max(d - c, Decimal("0"))
        sc = max(c - d, Decimal("0"))
        rows.append([code, t["libelle"], money(d), money(c),
                     money(sd) if sd else "", money(sc) if sc else ""])
    rows.append(["TOTAL", "", money(grand_d), money(grand_c), "", ""])

    periode = f"{date_debut or '—'} → {date_fin or '—'}"
    t_table = Table(rows, colWidths=[2*cm, 5.5*cm, 3*cm, 3*cm, 3*cm, 3*cm])
    ts2 = TableStyle(list(ts._cmds))
    ts2.add("FONTNAME", (0, len(rows)-1), (-1, len(rows)-1), "Helvetica-Bold")
    t_table.setStyle(ts2)

    elems = [
        Paragraph("Balance comptable", title_s),
        Paragraph(f"{residence.nom_residence} · Période : {periode}", sub_s),
        t_table,
    ]
    doc.build(elems)
    buf.seek(0)
    slug  = re.sub(r"[^a-z0-9]", "_", residence.nom_residence.lower())
    today = datetime.date.today().strftime("%Y%m%d")
    resp  = HttpResponse(buf.read(), content_type="application/pdf")
    resp["Content-Disposition"] = f'attachment; filename="balance_{slug}_{today}.pdf"'
    return resp


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def cpc_export_pdf(request):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from django.db.models import Sum

    residence = get_user_residence(request)
    if not residence:
        return Response({"detail": "Aucune résidence."}, status=400)
    date_debut = request.query_params.get("date_debut") or None
    date_fin   = request.query_params.get("date_fin")   or None
    D0 = Decimal("0")

    dep_qs = Depense.objects.select_related("compte").filter(residence=residence)
    rec_qs = Recette.objects.select_related("compte").filter(residence=residence)
    pai_qs = Paiement.objects.filter(residence=residence)
    if date_debut:
        dep_qs = dep_qs.filter(date_depense__gte=date_debut)
        rec_qs = rec_qs.filter(date_recette__gte=date_debut)
        pai_qs = pai_qs.filter(date_paiement__gte=date_debut)
    if date_fin:
        dep_qs = dep_qs.filter(date_depense__lte=date_fin)
        rec_qs = rec_qs.filter(date_recette__lte=date_fin)
        pai_qs = pai_qs.filter(date_paiement__lte=date_fin)

    charges_map = {}
    for d in dep_qs:
        key = (d.compte.code if d.compte else "6XX", d.compte.libelle if d.compte else "Divers")
        charges_map[key] = charges_map.get(key, D0) + (d.montant or D0)

    produits_map = {}
    for r in rec_qs:
        key = (r.compte.code if r.compte else "7XX", r.compte.libelle if r.compte else "Divers")
        produits_map[key] = produits_map.get(key, D0) + (r.montant or D0)
    total_pai = pai_qs.aggregate(t=Sum("montant"))["t"] or D0
    if total_pai:
        produits_map[(COMPTE_COPRO[0], "Charges récupérées — copropriétaires")] = \
            produits_map.get((COMPTE_COPRO[0], ""), D0) + total_pai

    total_charges  = sum(charges_map.values())
    total_produits = sum(produits_map.values())
    resultat       = total_produits - total_charges

    def money(v): return f"{float(v):,.2f} MAD"

    DARK  = colors.HexColor("#0F172A")
    RED   = colors.HexColor("#DC2626")
    GREEN = colors.HexColor("#059669")
    LIGHT = colors.HexColor("#F8FAFC")

    def section_table(header, rows_data, total_label, total_val, color):
        rows = [[header, ""]]
        for code, lib, m in rows_data:
            rows.append([f"{code} — {lib}", money(m)])
        rows.append([total_label, money(total_val)])
        t = Table(rows, colWidths=[11.5*cm, 4*cm])
        ts = TableStyle([
            ("BACKGROUND",    (0,0), (-1,0), color),
            ("TEXTCOLOR",     (0,0), (-1,0), colors.white),
            ("FONTNAME",      (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTNAME",      (0,-1), (-1,-1), "Helvetica-Bold"),
            ("FONTSIZE",      (0,0), (-1,-1), 9),
            ("ROWBACKGROUNDS",(0,1), (-1,-2), [LIGHT, colors.white]),
            ("GRID",          (0,0), (-1,-1), 0.4, colors.HexColor("#E2E8F0")),
            ("ALIGN",         (1,0), (1,-1), "RIGHT"),
            ("LEFTPADDING",   (0,0), (-1,-1), 6),
            ("RIGHTPADDING",  (0,0), (-1,-1), 6),
            ("TOPPADDING",    (0,0), (-1,-1), 4),
            ("BOTTOMPADDING", (0,0), (-1,-1), 4),
        ])
        t.setStyle(ts)
        return t

    styles = getSampleStyleSheet()
    title_s = ParagraphStyle("T", parent=styles["Normal"], fontSize=14, fontName="Helvetica-Bold", spaceAfter=4)
    sub_s   = ParagraphStyle("S", parent=styles["Normal"], fontSize=9, textColor=colors.HexColor("#64748B"), spaceAfter=12)
    res_s   = ParagraphStyle("R", parent=styles["Normal"], fontSize=11, fontName="Helvetica-Bold", spaceAfter=4)

    periode = f"{date_debut or '—'} → {date_fin or '—'}"
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
        leftMargin=2*cm, rightMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)

    ch_data  = [(k[0], k[1], v) for k, v in sorted(charges_map.items())]
    pr_data  = [(k[0], k[1], v) for k, v in sorted(produits_map.items())]

    sign = "+" if resultat >= 0 else ""
    color_res = GREEN if resultat >= 0 else RED

    elems = [
        Paragraph("Compte de Produits et Charges (CPC)", title_s),
        Paragraph(f"{residence.nom_residence} · Période : {periode}", sub_s),
        section_table("CHARGES", ch_data, "Total charges", total_charges, RED),
        Spacer(1, 12),
        section_table("PRODUITS", pr_data, "Total produits", total_produits, GREEN),
        Spacer(1, 16),
        Paragraph(
            f'<font color="{"#059669" if resultat >= 0 else "#DC2626"}">'
            f'Résultat net : {sign}{money(resultat)}</font>', res_s
        ),
    ]
    doc.build(elems)
    buf.seek(0)
    slug  = re.sub(r"[^a-z0-9]", "_", residence.nom_residence.lower())
    today = datetime.date.today().strftime("%Y%m%d")
    resp  = HttpResponse(buf.read(), content_type="application/pdf")
    resp["Content-Disposition"] = f'attachment; filename="cpc_{slug}_{today}.pdf"'
    return resp


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def bilan_export_pdf(request):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from django.db.models import Sum

    residence = get_user_residence(request)
    if not residence:
        return Response({"detail": "Aucune résidence."}, status=400)
    date_debut = request.query_params.get("date_debut") or None
    date_fin   = request.query_params.get("date_fin")   or None
    D0 = Decimal("0")

    dep_qs = Depense.objects.filter(residence=residence)
    rec_qs = Recette.objects.filter(residence=residence)
    pai_qs = Paiement.objects.filter(residence=residence)
    if date_debut:
        dep_qs = dep_qs.filter(date_depense__gte=date_debut)
        rec_qs = rec_qs.filter(date_recette__gte=date_debut)
        pai_qs = pai_qs.filter(date_paiement__gte=date_debut)
    if date_fin:
        dep_qs = dep_qs.filter(date_depense__lte=date_fin)
        rec_qs = rec_qs.filter(date_recette__lte=date_fin)
        pai_qs = pai_qs.filter(date_paiement__lte=date_fin)

    total_dep  = dep_qs.aggregate(t=Sum("montant"))["t"] or D0
    total_rec  = rec_qs.aggregate(t=Sum("montant"))["t"] or D0
    total_pai  = pai_qs.aggregate(t=Sum("montant"))["t"] or D0
    tresorerie = total_rec + total_pai - total_dep

    detail_qs = DetailAppelCharge.objects.filter(appel__residence=residence).exclude(statut="PAYE")
    if date_debut:
        detail_qs = detail_qs.filter(appel__date_emission__gte=date_debut)
    if date_fin:
        detail_qs = detail_qs.filter(appel__date_emission__lte=date_fin)
    agg = detail_qs.aggregate(du=Sum("montant"), recu=Sum("montant_recu"))
    creances = max((agg["du"] or D0) - (agg["recu"] or D0), D0)

    fond_qs = DetailAppelCharge.objects.filter(appel__residence=residence, appel__type_charge="FOND")
    if date_debut:
        fond_qs = fond_qs.filter(appel__date_emission__gte=date_debut)
    if date_fin:
        fond_qs = fond_qs.filter(appel__date_emission__lte=date_fin)
    fonds = fond_qs.aggregate(t=Sum("montant_recu"))["t"] or D0
    resultat = total_rec - total_dep

    def money(v): return f"{float(v):,.2f} MAD"
    DARK  = colors.HexColor("#0F172A")
    BLUE  = colors.HexColor("#1d4ed8")
    GREEN = colors.HexColor("#059669")
    LIGHT = colors.HexColor("#F8FAFC")

    styles = getSampleStyleSheet()
    title_s = ParagraphStyle("T", parent=styles["Normal"], fontSize=14, fontName="Helvetica-Bold", spaceAfter=4)
    sub_s   = ParagraphStyle("S", parent=styles["Normal"], fontSize=9, textColor=colors.HexColor("#64748B"), spaceAfter=12)

    actif_rows = [
        ["ACTIF", ""],
        ["Trésorerie nette (512)", money(tresorerie)],
        ["Créances copropriétaires (342)", money(creances)],
        ["TOTAL ACTIF", money(tresorerie + creances)],
    ]
    passif_rows = [
        ["PASSIF", ""],
        ["Résultat net de l'exercice", money(resultat)],
        ["Fonds travaux collectés", money(fonds)],
        ["Avances copropriétaires", money(total_pai)],
        ["TOTAL PASSIF", money(resultat + fonds + total_pai)],
    ]

    def make_table(rows, hdr_color):
        t = Table(rows, colWidths=[8.5*cm, 4*cm])
        ts = TableStyle([
            ("BACKGROUND",    (0,0), (-1,0), hdr_color),
            ("TEXTCOLOR",     (0,0), (-1,0), colors.white),
            ("FONTNAME",      (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTNAME",      (0,-1), (-1,-1), "Helvetica-Bold"),
            ("FONTSIZE",      (0,0), (-1,-1), 9),
            ("ROWBACKGROUNDS",(0,1), (-1,-2), [LIGHT, colors.white]),
            ("GRID",          (0,0), (-1,-1), 0.4, colors.HexColor("#E2E8F0")),
            ("ALIGN",         (1,0), (1,-1), "RIGHT"),
            ("LEFTPADDING",   (0,0), (-1,-1), 6),
            ("RIGHTPADDING",  (0,0), (-1,-1), 6),
            ("TOPPADDING",    (0,0), (-1,-1), 4),
            ("BOTTOMPADDING", (0,0), (-1,-1), 4),
        ])
        t.setStyle(ts)
        return t

    periode = f"{date_debut or '—'} → {date_fin or '—'}"
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
        leftMargin=2*cm, rightMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)

    elems = [
        Paragraph("Bilan simplifié", title_s),
        Paragraph(f"{residence.nom_residence} · Période : {periode}", sub_s),
        make_table(actif_rows, BLUE),
        Spacer(1, 12),
        make_table(passif_rows, GREEN),
    ]
    doc.build(elems)
    buf.seek(0)
    slug  = re.sub(r"[^a-z0-9]", "_", residence.nom_residence.lower())
    today = datetime.date.today().strftime("%Y%m%d")
    resp  = HttpResponse(buf.read(), content_type="application/pdf")
    resp["Content-Disposition"] = f'attachment; filename="bilan_{slug}_{today}.pdf"'
    return resp


# ============================================================
# Grand Livre — exports Excel & PDF
# ============================================================
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def grand_livre_export_excel(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from collections import OrderedDict

    residence = get_user_residence(request)
    if not residence:
        return Response({"detail": "Aucune résidence."}, status=400)
    date_debut    = request.query_params.get("date_debut") or None
    date_fin      = request.query_params.get("date_fin")   or None
    filter_compte = request.query_params.get("compte")     or None
    entries       = _build_entries(residence, date_debut, date_fin)

    thin   = Side(style="thin", color="E2E8F0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Group by compte
    comptes = OrderedDict()
    for e in entries:
        key = (e["compte_code"], e["compte_libelle"])
        if key not in comptes:
            comptes[key] = []
        comptes[key].append(e)

    wb = Workbook()
    ws = wb.active
    ws.title = "Grand Livre"
    for col, w in zip("ABCDEFGH", [12, 14, 14, 36, 14, 14, 14, 14]):
        ws.column_dimensions[col].width = w

    row = 1
    ws.cell(row, 1, "GRAND LIVRE").font = Font(bold=True, size=13)
    ws.cell(row, 3, f"Période : {date_debut or '—'} → {date_fin or '—'}").font = Font(size=9, color="64748B")
    row += 2

    for (code, libelle), lines in sorted(comptes.items()):
        if filter_compte and filter_compte.lower() not in code.lower() and filter_compte.lower() not in libelle.lower():
            continue
        # Compte header
        for col in range(1, 9):
            c = ws.cell(row, col)
            c.fill = PatternFill("solid", fgColor="0F172A")
            c.font = Font(bold=True, color="FFFFFF", size=10)
        ws.cell(row, 1, code)
        ws.cell(row, 2, libelle).font = Font(bold=True, color="FFFFFF", size=10)
        row += 1
        # Column headers
        for col, h in enumerate(["Date", "Type", "Pièce", "Libellé", "Débit", "Crédit", "Solde"], 1):
            c = ws.cell(row, col, h)
            c.fill = PatternFill("solid", fgColor="334155")
            c.font = Font(bold=True, color="FFFFFF", size=9)
            c.border = border
        row += 1
        # Entries
        total_d = Decimal("0")
        total_c = Decimal("0")
        for e in lines:
            total_d += e["debit"]
            total_c += e["credit"]
            for col, val in enumerate([
                str(e["date"]), e["type"], e["piece"] or "",
                (e["libelle"] or "")[:50],
                float(e["debit"]) or None, float(e["credit"]) or None,
                float(total_d - total_c),
            ], 1):
                c = ws.cell(row, col, val)
                c.border = border
                if col in (5, 6, 7) and val is not None:
                    c.number_format = "#,##0.00"
            row += 1
        # Compte total
        ws.cell(row, 4, "Solde du compte").font = Font(bold=True)
        c5 = ws.cell(row, 5, float(total_d)); c5.number_format = "#,##0.00"; c5.font = Font(bold=True)
        c6 = ws.cell(row, 6, float(total_c)); c6.number_format = "#,##0.00"; c6.font = Font(bold=True)
        c7 = ws.cell(row, 7, float(total_d - total_c)); c7.number_format = "#,##0.00"; c7.font = Font(bold=True)
        row += 2

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    slug  = re.sub(r"[^a-z0-9]", "_", residence.nom_residence.lower())
    today = datetime.date.today().strftime("%Y%m%d")
    resp  = HttpResponse(buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = f'attachment; filename="grand_livre_{slug}_{today}.xlsx"'
    return resp


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def grand_livre_export_pdf(request):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from collections import OrderedDict

    residence = get_user_residence(request)
    if not residence:
        return Response({"detail": "Aucune résidence."}, status=400)
    date_debut    = request.query_params.get("date_debut") or None
    date_fin      = request.query_params.get("date_fin")   or None
    filter_compte = request.query_params.get("compte")     or None
    entries       = _build_entries(residence, date_debut, date_fin)

    # Group by compte
    comptes = OrderedDict()
    for e in entries:
        key = (e["compte_code"], e["compte_libelle"])
        if key not in comptes:
            comptes[key] = []
        comptes[key].append(e)

    DARK  = colors.HexColor("#0F172A")
    MID   = colors.HexColor("#334155")
    LIGHT = colors.HexColor("#F8FAFC")

    def money(v):
        fv = float(v)
        return f"{fv:,.2f}" if fv else ""

    styles = getSampleStyleSheet()
    title_s = ParagraphStyle("T", parent=styles["Normal"], fontSize=14, fontName="Helvetica-Bold", spaceAfter=4)
    sub_s   = ParagraphStyle("S", parent=styles["Normal"], fontSize=9, textColor=colors.HexColor("#64748B"), spaceAfter=8)
    acct_s  = ParagraphStyle("A", parent=styles["Normal"], fontSize=10, fontName="Helvetica-Bold", spaceAfter=2)

    periode = f"{date_debut or '—'} → {date_fin or '—'}"
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
        leftMargin=1.5*cm, rightMargin=1.5*cm, topMargin=1.8*cm, bottomMargin=1.8*cm)

    entry_ts = TableStyle([
        ("BACKGROUND",    (0, 0), (-1, 0), MID),
        ("TEXTCOLOR",     (0, 0), (-1, 0), colors.white),
        ("FONTNAME",      (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",      (0, 0), (-1, -1), 8),
        ("ROWBACKGROUNDS",(0, 1), (-1, -1), [LIGHT, colors.white]),
        ("GRID",          (0, 0), (-1, -1), 0.4, colors.HexColor("#E2E8F0")),
        ("ALIGN",         (4, 0), (-1, -1), "RIGHT"),
        ("LEFTPADDING",   (0, 0), (-1, -1), 5),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 5),
        ("TOPPADDING",    (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ])

    elems = [
        Paragraph("Grand Livre", title_s),
        Paragraph(f"{residence.nom_residence} · Période : {periode}", sub_s),
    ]

    for (code, libelle), lines in sorted(comptes.items()):
        if filter_compte and filter_compte.lower() not in code.lower() and filter_compte.lower() not in libelle.lower():
            continue
        if len(lines) > 100:
            lines = lines[:100]

        total_d = Decimal("0")
        total_c = Decimal("0")
        rows = [["Date", "Type", "Pièce", "Libellé", "Débit", "Crédit", "Solde"]]
        for e in lines:
            total_d += e["debit"]
            total_c += e["credit"]
            rows.append([
                str(e["date"]), e["type"], (e["piece"] or "")[:10],
                (e["libelle"] or "")[:38],
                money(e["debit"]), money(e["credit"]),
                money(total_d - total_c),
            ])
        rows.append(["", "", "", "Solde",
                     money(total_d), money(total_c), money(total_d - total_c)])

        ts2 = TableStyle(list(entry_ts._cmds))
        ts2.add("FONTNAME", (0, len(rows)-1), (-1, len(rows)-1), "Helvetica-Bold")
        t = Table(rows, colWidths=[2.2*cm, 2*cm, 2.2*cm, 6*cm, 2.6*cm, 2.6*cm, 2.6*cm])
        t.setStyle(ts2)

        elems.append(Paragraph(f"{code} — {libelle}", acct_s))
        elems.append(t)
        elems.append(Spacer(1, 10))

    doc.build(elems)
    buf.seek(0)
    slug  = re.sub(r"[^a-z0-9]", "_", residence.nom_residence.lower())
    today = datetime.date.today().strftime("%Y%m%d")
    resp  = HttpResponse(buf.read(), content_type="application/pdf")
    resp["Content-Disposition"] = f'attachment; filename="grand_livre_{slug}_{today}.pdf"'
    return resp


# ============================================================
# Années disponibles (pour le sélecteur d'exercice)
# ============================================================
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def comptabilite_annees(request):
    """Retourne les années pour lesquelles des dépenses ou recettes existent."""
    residence = get_user_residence(request)
    if not residence:
        return Response({"annees": []})
    from django.db.models.functions import ExtractYear
    dep_years = (
        Depense.objects.filter(residence=residence)
        .annotate(y=ExtractYear("date_depense"))
        .values_list("y", flat=True)
        .distinct()
    )
    rec_years = (
        Recette.objects.filter(residence=residence)
        .annotate(y=ExtractYear("date_recette"))
        .values_list("y", flat=True)
        .distinct()
    )
    years = sorted({y for y in list(dep_years) + list(rec_years) if y}, reverse=True)
    return Response({"annees": years})
