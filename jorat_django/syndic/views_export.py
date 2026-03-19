"""
Export des données — génère un fichier Excel (xlsx) multi-feuilles
filtré par résidence de l'utilisateur connecté.
"""
import io
import datetime

from django.http import HttpResponse
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from .models import (
    Residence, Lot, Personne, Groupe,
    AppelCharge, DetailAppelCharge, AffectationPaiement,
    Paiement, CaisseMouvement, Depense, Recette,
    BureauSyndical, AssembleeGenerale, Resolution, DocumentGouvernance,
)
from .views import get_user_residence


# ── helpers ────────────────────────────────────────────────────────────────

HEADER_FILL = PatternFill("solid", fgColor="1E40AF")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")


def _header(ws, columns):
    ws.append(columns)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN


def _auto_width(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)


# ── sheets ─────────────────────────────────────────────────────────────────

def _sheet_residence(wb, residence):
    ws = wb.create_sheet("Résidence")
    _header(ws, ["ID", "Nom", "Ville", "Adresse", "Mode appel", "Classification", "Statut"])
    ws.append([
        residence.id,
        residence.nom_residence,
        residence.ville_residence,
        residence.adresse_residence or "",
        residence.get_mode_appel_charge_display(),
        residence.get_classification_display(),
        residence.get_statut_residence_display(),
    ])
    _auto_width(ws)


def _sheet_groupes(wb, residence):
    ws = wb.create_sheet("Groupes")
    _header(ws, ["ID", "Nom groupe", "Description"])
    for g in Groupe.objects.filter(residence=residence):
        ws.append([g.id, g.nom_groupe, g.description])
    _auto_width(ws)


def _sheet_personnes(wb, residence):
    ws = wb.create_sheet("Personnes")
    _header(ws, ["ID", "Nom", "Prénom", "CIN", "Téléphone", "Email", "Type"])
    for p in Personne.objects.filter(residence=residence):
        ws.append([p.id, p.nom, p.prenom, p.cin or "", p.telephone, p.email, p.get_type_personne_display()])
    _auto_width(ws)


def _sheet_lots(wb, residence):
    ws = wb.create_sheet("Lots")
    _header(ws, ["ID", "Numéro", "Type", "Étage", "Surface", "Montant réf.", "Statut", "Groupe", "Propriétaire"])
    for lot in Lot.objects.select_related("groupe", "proprietaire").filter(residence=residence):
        ws.append([
            lot.id,
            lot.numero_lot,
            lot.get_type_lot_display(),
            lot.etage_lot,
            float(lot.surface_lot) if lot.surface_lot else "",
            float(lot.montant_ref),
            lot.get_statut_lot_display(),
            lot.groupe.nom_groupe if lot.groupe else "",
            f"{lot.proprietaire.nom} {lot.proprietaire.prenom}".strip() if lot.proprietaire else "",
        ])
    _auto_width(ws)


def _sheet_appels(wb, residence):
    ws = wb.create_sheet("Appels de charge")
    _header(ws, ["ID", "Type", "Exercice", "Période", "Code", "Nom fond", "Date émission"])
    for a in AppelCharge.objects.filter(residence=residence):
        ws.append([
            a.id,
            a.get_type_charge_display(),
            a.exercice,
            a.get_periode_display(),
            a.code_fond or "",
            a.nom_fond or "",
            str(a.date_emission),
        ])
    _auto_width(ws)


def _sheet_details_appels(wb, residence):
    ws = wb.create_sheet("Détails appels")
    _header(ws, ["ID", "Appel ID", "Lot", "Montant dû", "Montant reçu", "Statut"])
    for d in DetailAppelCharge.objects.select_related("appel", "lot").filter(appel__residence=residence):
        ws.append([
            d.id,
            d.appel_id,
            d.lot.numero_lot,
            float(d.montant),
            float(d.montant_recu),
            d.get_statut_display(),
        ])
    _auto_width(ws)


def _sheet_paiements(wb, residence):
    ws = wb.create_sheet("Paiements")
    _header(ws, ["ID", "Date", "Lot", "Montant", "Référence", "Période comptable", "Mode paiement"])
    for p in Paiement.objects.select_related("lot").filter(residence=residence):
        ws.append([
            p.id, str(p.date_paiement), p.lot.numero_lot, float(p.montant), p.reference,
            p.mois or "", p.mode_paiement or "",
        ])
    _auto_width(ws)


def _sheet_affectations(wb, residence):
    ws = wb.create_sheet("Affectations")
    _header(ws, ["ID", "Paiement ID", "Détail ID", "Lot", "Montant affecté"])
    for a in AffectationPaiement.objects.select_related("paiement__lot", "detail").filter(paiement__residence=residence):
        ws.append([
            a.id,
            a.paiement_id,
            a.detail_id or "",
            a.paiement.lot.numero_lot,
            float(a.montant_affecte),
        ])
    _auto_width(ws)


def _sheet_caisse(wb, residence):
    ws = wb.create_sheet("Caisse")
    _header(ws, ["ID", "Date", "Type", "Sens", "Montant", "Libellé", "Commentaire"])
    for m in CaisseMouvement.objects.filter(residence=residence):
        ws.append([
            m.id,
            str(m.date_mouvement),
            m.get_type_mouvement_display(),
            m.get_sens_display(),
            float(m.montant),
            m.libelle,
            m.commentaire,
        ])
    _auto_width(ws)


def _sheet_depenses(wb, residence):
    ws = wb.create_sheet("Dépenses")
    _header(ws, ["ID", "Date", "Libellé", "Montant", "Compte", "Catégorie", "Fournisseur", "Réf. facture", "Période comptable"])
    for d in Depense.objects.select_related("compte", "categorie", "fournisseur").filter(residence=residence):
        ws.append([
            d.id,
            str(d.date_depense),
            d.libelle,
            float(d.montant),
            str(d.compte) if d.compte else "",
            str(d.categorie) if d.categorie else "",
            str(d.fournisseur) if d.fournisseur else "",
            d.facture_reference,
            d.mois or "",
        ])
    _auto_width(ws)


def _sheet_recettes(wb, residence):
    ws = wb.create_sheet("Recettes")
    _header(ws, ["ID", "Date", "Libellé", "Montant", "Compte", "Source", "Commentaire", "Période comptable"])
    for r in Recette.objects.select_related("compte").filter(residence=residence):
        ws.append([
            r.id,
            str(r.date_recette),
            r.libelle,
            float(r.montant),
            str(r.compte) if r.compte else "",
            r.source,
            r.commentaire,
            r.mois or "",
        ])
    _auto_width(ws)


def _sheet_bureau(wb, residence):
    ws = wb.create_sheet("Bureau Syndical")
    _header(ws, ["ID", "Personne", "Fonction", "Date début", "Date fin", "Actif"])
    for b in BureauSyndical.objects.select_related("personne").filter(residence=residence):
        ws.append([
            b.id,
            f"{b.personne.nom} {b.personne.prenom}".strip(),
            b.get_fonction_display(),
            str(b.date_debut),
            str(b.date_fin) if b.date_fin else "",
            "Oui" if b.actif else "Non",
        ])
    _auto_width(ws)


def _sheet_assemblees(wb, residence):
    ws = wb.create_sheet("Assemblées Générales")
    _header(ws, ["ID", "Date", "Type", "Statut", "Ordre du jour"])
    for a in AssembleeGenerale.objects.filter(residence=residence):
        ws.append([
            a.id,
            str(a.date_ag),
            a.get_type_ag_display(),
            a.get_statut_display(),
            a.ordre_du_jour,
        ])
    _auto_width(ws)


def _sheet_resolutions(wb, residence):
    ws = wb.create_sheet("Résolutions")
    _header(ws, ["ID", "AG ID", "N°", "Titre", "Pour", "Contre", "Abstentions", "Résultat"])
    for r in Resolution.objects.select_related("assemblee_generale").filter(assemblee_generale__residence=residence):
        ws.append([
            r.id,
            r.assemblee_generale_id,
            r.numero,
            r.titre,
            r.voix_pour,
            r.voix_contre,
            r.abstention,
            r.get_resultat_display(),
        ])
    _auto_width(ws)


def _sheet_documents(wb, residence):
    ws = wb.create_sheet("Documents")
    _header(ws, ["ID", "Titre", "Type", "Date", "Visible résident"])
    for d in DocumentGouvernance.objects.filter(residence=residence):
        ws.append([
            d.id,
            d.titre,
            d.get_type_document_display(),
            str(d.date),
            "Oui" if d.visible_resident else "Non",
        ])
    _auto_width(ws)


# ── mapping key -> builder ──────────────────────────────────────────────────

SHEETS = {
    "residence":       _sheet_residence,
    "groupes":         _sheet_groupes,
    "personnes":       _sheet_personnes,
    "lots":            _sheet_lots,
    "appels":          _sheet_appels,
    "details-appels":  _sheet_details_appels,
    "paiements":       _sheet_paiements,
    "affectations":    _sheet_affectations,
    "caisse":          _sheet_caisse,
    "depenses":        _sheet_depenses,
    "recettes":        _sheet_recettes,
    "bureau":          _sheet_bureau,
    "assemblees":      _sheet_assemblees,
    "resolutions":     _sheet_resolutions,
    "documents":       _sheet_documents,
}

ALL_KEYS = list(SHEETS.keys())


# ── view ───────────────────────────────────────────────────────────────────

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def export_excel(request):
    residence = get_user_residence(request)
    if not residence:
        from rest_framework.response import Response
        return Response({"detail": "Aucune résidence associée."}, status=400)

    full = request.GET.get("full", "false").lower() == "true"
    if full:
        keys = ALL_KEYS
    else:
        raw = request.GET.get("datasets", "")
        keys = [k.strip() for k in raw.split(",") if k.strip() in SHEETS]
        if not keys:
            keys = ALL_KEYS

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    for key in keys:
        SHEETS[key](wb, residence)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    date_str = datetime.date.today().strftime("%Y%m%d")
    nom = residence.nom_residence.replace(" ", "_")
    filename = f"export_{nom}_{date_str}.xlsx"

    response = HttpResponse(
        buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response
