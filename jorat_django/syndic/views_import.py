"""
Import des données depuis un fichier Excel (.xlsx)
Endpoint: POST /api/import/excel/    — validate ou import
           GET  /api/import/template/ — télécharger un modèle
"""
import io
import datetime
from decimal import Decimal, InvalidOperation

from django.http import HttpResponse
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from .models import (
    Lot, Personne, Groupe, CompteComptable,
    AppelCharge, DetailAppelCharge, Paiement, AffectationPaiement,
    Recette, Depense, CategorieDepense, Fournisseur,
)
from .views import get_user_residence


# ── Column specs ─────────────────────────────────────────────────────────────

SPECS = {
    # ── Référentiel ──────────────────────────────────────────────────────────
    "lots": {
        "columns":  ["numero_lot", "type_lot", "etage_lot", "surface", "montant_ref", "groupe", "proprietaire"],
        "required": ["numero_lot"],
        "description": "Lots (appartements, bureaux, etc.)",
    },
    "personnes": {
        "columns":  ["nom", "prenom", "telephone", "email", "type_personne"],
        "required": ["nom"],
        "description": "Personnes (propriétaires, occupants)",
    },
    "plan-comptable": {
        "columns":  ["code_compte", "libelle", "type_compte"],
        "required": ["code_compte", "libelle"],
        "description": "Plan comptable",
    },
    # ── Finance ──────────────────────────────────────────────────────────────
    "appel-charge": {
        "columns":  ["type_charge", "exercice", "periode", "nom_fond", "description_appel", "date_emission"],
        "required": ["type_charge", "exercice"],
        "description": "Appels de charge",
        "note": "type_charge: CHARGE ou FOND. periode: JAN, FEV, MAR, AVR, MAI, JUI, JUL, AOU, SEP, OCT, NOV, DEC, ANNEE",
    },
    "detail-appel": {
        "columns":  ["code_appel", "numero_lot", "montant"],
        "required": ["code_appel", "numero_lot", "montant"],
        "description": "Détails appels de charge",
        "note": "code_appel: code_fond de l'appel (ex: JAN_2024 ou AF_2024 1)",
    },
    "paiements": {
        "columns":  ["numero_lot", "date_paiement", "montant", "reference", "mois", "mode_paiement"],
        "required": ["numero_lot", "montant"],
        "description": "Paiements copropriétaires",
        "note": "Mise à jour automatique de la caisse. mois: JAN-DEC. mode_paiement: ESPECES, VIREMENT ou CHEQUE.",
    },
    "affectation-paiement": {
        "columns":  ["reference_paiement", "numero_lot", "code_appel", "montant_affecte"],
        "required": ["reference_paiement", "numero_lot", "code_appel", "montant_affecte"],
        "description": "Affectations paiements → appels",
        "note": "reference_paiement: référence du paiement. code_appel: code_fond de l'appel.",
    },
    "recettes": {
        "columns":  ["date_recette", "montant", "libelle", "code_compte", "source", "mois"],
        "required": ["date_recette", "montant", "libelle", "code_compte"],
        "description": "Recettes",
        "note": "Mise à jour automatique de la caisse. code_compte: code du compte comptable. mois: JAN-DEC.",
    },
    "depenses": {
        "columns":  ["date_depense", "montant", "libelle", "code_compte", "categorie", "fournisseur", "facture_reference", "mois"],
        "required": ["date_depense", "montant", "libelle", "code_compte"],
        "description": "Dépenses",
        "note": "Mise à jour automatique de la caisse. categorie et fournisseur sont optionnels. mois: JAN-DEC.",
    },
}

TYPE_LOT_VALUES    = ["APPARTEMENT", "VILLA", "MAISON", "LOCAL", "COMMERCE", "BUREAU", "AUTRE"]
TYPE_PERS_VALUES   = ["PHYSIQUE", "MORALE"]
TYPE_COMPTE_VALUES = ["CHARGE", "PRODUIT", "TRESORERIE"]
TYPE_COMPTE_MAP    = {
    "charge":     "CHARGE",
    "produit":    "PRODUIT",
    "tresorerie": "TRESORERIE",
    "actif":      "TRESORERIE",
    "passif":     "TRESORERIE",
}
PERIODE_VALUES       = ["ANNEE", "JAN", "FEV", "MAR", "AVR", "MAI", "JUI", "JUL", "AOU", "SEP", "OCT", "NOV", "DEC", "FOND"]
MOIS_VALUES          = ["JAN", "FEV", "MAR", "AVR", "MAI", "JUN", "JUL", "AOU", "SEP", "OCT", "NOV", "DEC"]
MODE_PAIEMENT_VALUES = ["ESPECES", "VIREMENT", "CHEQUE"]


# ── Helpers ──────────────────────────────────────────────────────────────────

def _dec(val, field_name, errors, row_num, min_val=Decimal("0")):
    """Parse decimal, append error if invalid. Returns Decimal or None."""
    raw = str(val or "").strip().replace(",", ".")
    if not raw:
        return None
    try:
        d = Decimal(raw)
        if d < min_val:
            raise ValueError
        return d
    except (InvalidOperation, ValueError):
        errors.append({"row": row_num, "message": f"{field_name} invalide: '{raw}'"})
        return None


def _parse_date(val):
    """Parse date from string, datetime, or Excel float. Returns date or None."""
    if val is None:
        return None
    if isinstance(val, (datetime.datetime,)):
        return val.date()
    if isinstance(val, datetime.date):
        return val
    s = str(val).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


# ── Template ─────────────────────────────────────────────────────────────────

HEADER_FILL = PatternFill("solid", fgColor="1E40AF")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)

EXAMPLE_ROWS = {
    "lots": [
        ["A101", "APPARTEMENT", "1", "85.5", "500", "Bâtiment A", "DUPONT Jean"],
        ["B202", "BUREAU",      "2", "42.0", "300", "",           ""],
    ],
    "personnes": [
        ["MARTIN", "Sophie", "0612345678", "sophie@mail.com", "PHYSIQUE"],
        ["ACME SA", "",      "0522334455", "contact@acme.ma", "MORALE"],
    ],
    "plan-comptable": [
        ["606", "Fournitures d'entretien", "CHARGE"],
        ["701", "Appels de charges",       "PRODUIT"],
        ["512", "Banque",                  "TRESORERIE"],
    ],
    "appel-charge": [
        ["CHARGE", 2024, "JAN", "", "Appel charges janvier", "2024-01-01"],
        ["FOND",   2024, "",    "Ravalement façade", "Travaux façade", "2024-03-15"],
    ],
    "detail-appel": [
        ["JAN_2024", "A101", "500.00"],
        ["JAN_2024", "B202", "300.00"],
    ],
    "paiements": [
        ["A101", "2024-01-15", "500.00", "VIR-2024-001", "JAN", "VIREMENT"],
        ["B202", "2024-01-20", "300.00", "VIR-2024-002", "JAN", "CHEQUE"],
    ],
    "affectation-paiement": [
        ["VIR-2024-001", "A101", "JAN_2024", "500.00"],
    ],
    "recettes": [
        ["2024-01-31", "1200.00", "Subvention commune", "701", "Mairie", "JAN"],
    ],
    "depenses": [
        ["2024-01-10", "450.00", "Nettoyage parties communes", "606", "Nettoyage", "Entreprise Netto", "FAC-001", "JAN"],
    ],
}


def _build_template(dataset):
    spec = SPECS[dataset]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = spec["description"][:31]
    ws.append(spec["columns"])
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    for row in EXAMPLE_ROWS.get(dataset, []):
        ws.append(row)
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── Validators ────────────────────────────────────────────────────────────────

def _validate_lots(rows, residence):
    groupes   = {g.nom_groupe.strip().lower(): g for g in Groupe.objects.filter(residence=residence)}
    personnes = {f"{p.nom.strip()} {p.prenom.strip()}".strip().lower(): p
                 for p in Personne.objects.filter(residence=residence)}
    existing_lots = set(Lot.objects.filter(residence=residence).values_list("numero_lot", flat=True))

    valid, errors = [], []
    for i, row in enumerate(rows, start=2):
        numero = str(row.get("numero_lot", "")).strip()
        if not numero:
            errors.append({"row": i, "message": "numero_lot est obligatoire"}); continue
        if numero in existing_lots:
            errors.append({"row": i, "message": f"Le lot '{numero}' existe déjà"}); continue

        type_lot = str(row.get("type_lot", "APPARTEMENT")).strip().upper() or "APPARTEMENT"
        if type_lot not in TYPE_LOT_VALUES:
            type_lot = "AUTRE"

        errs_before = len(errors)
        surface     = _dec(row.get("surface"),     "surface",     errors, i)
        montant_ref = _dec(row.get("montant_ref"), "montant_ref", errors, i) or Decimal("0")
        if len(errors) > errs_before:
            continue

        groupe_obj = None
        raw_grp = str(row.get("groupe", "")).strip()
        if raw_grp:
            groupe_obj = groupes.get(raw_grp.lower())
            if not groupe_obj:
                errors.append({"row": i, "message": f"Groupe introuvable: '{raw_grp}'"}); continue

        proprio_obj = None
        raw_prop = str(row.get("proprietaire", "")).strip()
        if raw_prop:
            proprio_obj = personnes.get(raw_prop.lower())
            if not proprio_obj:
                errors.append({"row": i, "message": f"Propriétaire introuvable: '{raw_prop}'"}); continue

        valid.append({
            "numero_lot": numero, "type_lot": type_lot,
            "etage_lot":  str(row.get("etage_lot", "")).strip(),
            "surface_lot": surface, "montant_ref": montant_ref,
            "groupe": groupe_obj, "proprietaire": proprio_obj,
        })
    return valid, errors


def _validate_personnes(rows, residence):
    valid, errors = [], []
    for i, row in enumerate(rows, start=2):
        nom = str(row.get("nom", "")).strip()
        if not nom:
            errors.append({"row": i, "message": "nom est obligatoire"}); continue
        type_p = str(row.get("type_personne", "PHYSIQUE")).strip().upper() or "PHYSIQUE"
        if type_p not in TYPE_PERS_VALUES:
            type_p = "PHYSIQUE"
        valid.append({
            "nom": nom, "prenom": str(row.get("prenom", "")).strip(),
            "telephone": str(row.get("telephone", "")).strip(),
            "email": str(row.get("email", "")).strip(), "type_personne": type_p,
        })
    return valid, errors


def _validate_plan_comptable(rows, residence):
    existing_codes = set(CompteComptable.objects.filter(residence=residence).values_list("code", flat=True))
    valid, errors, seen_codes = [], [], set()
    for i, row in enumerate(rows, start=2):
        code = str(row.get("code_compte", "")).strip()
        if not code:
            errors.append({"row": i, "message": "code_compte est obligatoire"}); continue
        libelle = str(row.get("libelle", "")).strip()
        if not libelle:
            errors.append({"row": i, "message": "libelle est obligatoire"}); continue
        if code in existing_codes:
            errors.append({"row": i, "message": f"Code '{code}' existe déjà"}); continue
        if code in seen_codes:
            errors.append({"row": i, "message": f"Code '{code}' en double dans le fichier"}); continue
        raw_type = str(row.get("type_compte", "")).strip().lower()
        type_compte = TYPE_COMPTE_MAP.get(raw_type) or (raw_type.upper() if raw_type.upper() in TYPE_COMPTE_VALUES else None)
        if not type_compte:
            errors.append({"row": i, "message": f"type_compte invalide: '{raw_type}'"}); continue
        seen_codes.add(code)
        valid.append({"code": code, "libelle": libelle, "type_compte": type_compte})
    return valid, errors


def _validate_appel_charge(rows, residence):
    # Pre-load existing (exercice, periode) for CHARGE uniqueness check
    existing_charge = set(
        AppelCharge.objects.filter(residence=residence, type_charge="CHARGE")
        .values_list("exercice", "periode")
    )
    valid, errors, seen = [], [], set()
    for i, row in enumerate(rows, start=2):
        type_charge = str(row.get("type_charge", "")).strip().upper()
        if type_charge not in ("CHARGE", "FOND"):
            errors.append({"row": i, "message": "type_charge doit être CHARGE ou FOND"}); continue

        raw_ex = str(row.get("exercice", "")).strip()
        try:
            exercice = int(float(raw_ex)) if raw_ex else None
            if not exercice or exercice < 2000 or exercice > 2050:
                raise ValueError
        except (ValueError, TypeError):
            errors.append({"row": i, "message": f"exercice invalide: '{raw_ex}'"}); continue

        periode = str(row.get("periode", "")).strip().upper() or ("FOND" if type_charge == "FOND" else "")
        if type_charge == "CHARGE":
            if not periode or periode not in PERIODE_VALUES or periode == "FOND":
                errors.append({"row": i, "message": f"periode invalide pour CHARGE: '{periode}'. Valeurs: JAN…DEC, ANNEE"}); continue
            key = (exercice, periode)
            if key in existing_charge or key in seen:
                errors.append({"row": i, "message": f"Appel CHARGE {exercice}/{periode} existe déjà"}); continue
            seen.add(key)
        else:
            periode = "FOND"

        date_emission = _parse_date(row.get("date_emission")) or datetime.date.today()

        valid.append({
            "type_charge": type_charge, "exercice": exercice, "periode": periode,
            "nom_fond": str(row.get("nom_fond", "")).strip(),
            "description_appel": str(row.get("description_appel", "")).strip(),
            "date_emission": date_emission,
        })
    return valid, errors


def _validate_detail_appel(rows, residence):
    # Pre-load appels and lots
    appels = {a.code_fond: a for a in AppelCharge.objects.filter(residence=residence) if a.code_fond}
    lots   = {l.numero_lot: l for l in Lot.objects.filter(residence=residence)}
    existing_details = set(
        DetailAppelCharge.objects.filter(appel__residence=residence)
        .values_list("appel__code_fond", "lot__numero_lot")
    )
    valid, errors, seen = [], [], set()
    for i, row in enumerate(rows, start=2):
        code_appel = str(row.get("code_appel", "")).strip()
        numero_lot = str(row.get("numero_lot", "")).strip()
        if not code_appel:
            errors.append({"row": i, "message": "code_appel est obligatoire"}); continue
        if not numero_lot:
            errors.append({"row": i, "message": "numero_lot est obligatoire"}); continue

        appel_obj = appels.get(code_appel)
        if not appel_obj:
            errors.append({"row": i, "message": f"Appel introuvable avec code '{code_appel}'"}); continue
        lot_obj = lots.get(numero_lot)
        if not lot_obj:
            errors.append({"row": i, "message": f"Lot introuvable: '{numero_lot}'"}); continue

        key = (code_appel, numero_lot)
        if key in existing_details or key in seen:
            errors.append({"row": i, "message": f"Détail ({code_appel}, {numero_lot}) existe déjà"}); continue

        errs_before = len(errors)
        montant = _dec(row.get("montant"), "montant", errors, i, min_val=Decimal("0.01"))
        if len(errors) > errs_before:
            continue
        seen.add(key)
        valid.append({"appel": appel_obj, "lot": lot_obj, "montant": montant})
    return valid, errors


def _validate_paiements(rows, residence):
    lots = {l.numero_lot: l for l in Lot.objects.filter(residence=residence)}
    valid, errors = [], []
    for i, row in enumerate(rows, start=2):
        numero_lot = str(row.get("numero_lot", "")).strip()
        if not numero_lot:
            errors.append({"row": i, "message": "numero_lot est obligatoire"}); continue
        lot_obj = lots.get(numero_lot)
        if not lot_obj:
            errors.append({"row": i, "message": f"Lot introuvable: '{numero_lot}'"}); continue

        errs_before = len(errors)
        montant = _dec(row.get("montant"), "montant", errors, i, min_val=Decimal("0.01"))
        if len(errors) > errs_before:
            continue

        date_paiement = _parse_date(row.get("date_paiement")) or datetime.date.today()

        raw_mois = str(row.get("mois", "")).strip().upper()
        mois = raw_mois if raw_mois in MOIS_VALUES else None

        raw_mode = str(row.get("mode_paiement", "")).strip().upper()
        mode_paiement = raw_mode if raw_mode in MODE_PAIEMENT_VALUES else None

        valid.append({
            "lot": lot_obj, "date_paiement": date_paiement,
            "montant": montant, "reference": str(row.get("reference", "")).strip(),
            "mois": mois, "mode_paiement": mode_paiement,
        })
    return valid, errors


def _validate_affectation_paiement(rows, residence):
    lots = {l.numero_lot: l for l in Lot.objects.filter(residence=residence)}
    # Map paiements by (reference, lot_id)
    paiements_by_ref = {}
    for p in Paiement.objects.filter(residence=residence).select_related("lot"):
        key = (p.reference.strip(), p.lot.numero_lot)
        paiements_by_ref.setdefault(key, []).append(p)

    # Map details by (code_fond, numero_lot)
    details_map = {}
    for d in (DetailAppelCharge.objects
              .filter(appel__residence=residence)
              .select_related("appel", "lot")):
        if d.appel.code_fond:
            details_map[(d.appel.code_fond, d.lot.numero_lot)] = d

    existing = set(
        AffectationPaiement.objects
        .filter(paiement__residence=residence)
        .values_list("paiement_id", "detail_id")
    )
    valid, errors = [], []
    for i, row in enumerate(rows, start=2):
        ref_pmt    = str(row.get("reference_paiement", "")).strip()
        numero_lot = str(row.get("numero_lot", "")).strip()
        code_appel = str(row.get("code_appel", "")).strip()
        if not ref_pmt:
            errors.append({"row": i, "message": "reference_paiement est obligatoire"}); continue
        if not numero_lot:
            errors.append({"row": i, "message": "numero_lot est obligatoire"}); continue
        if not code_appel:
            errors.append({"row": i, "message": "code_appel est obligatoire"}); continue

        pmts = paiements_by_ref.get((ref_pmt, numero_lot), [])
        if not pmts:
            errors.append({"row": i, "message": f"Paiement introuvable (référence='{ref_pmt}', lot='{numero_lot}')"}); continue
        pmt = pmts[0]  # use most recent if duplicates

        detail = details_map.get((code_appel, numero_lot))
        if not detail:
            errors.append({"row": i, "message": f"Détail introuvable ({code_appel}, lot {numero_lot})"}); continue

        if (pmt.id, detail.id) in existing:
            errors.append({"row": i, "message": f"Affectation ({ref_pmt} → {code_appel}) existe déjà"}); continue

        errs_before = len(errors)
        montant_affecte = _dec(row.get("montant_affecte"), "montant_affecte", errors, i, min_val=Decimal("0.01"))
        if len(errors) > errs_before:
            continue

        valid.append({"paiement": pmt, "detail": detail, "montant_affecte": montant_affecte})
    return valid, errors


def _validate_recettes(rows, residence):
    comptes = {c.code: c for c in CompteComptable.objects.filter(residence=residence)}
    valid, errors = [], []
    for i, row in enumerate(rows, start=2):
        libelle = str(row.get("libelle", "")).strip()
        if not libelle:
            errors.append({"row": i, "message": "libelle est obligatoire"}); continue

        code_compte = str(row.get("code_compte", "")).strip()
        if not code_compte:
            errors.append({"row": i, "message": "code_compte est obligatoire"}); continue
        compte_obj = comptes.get(code_compte)
        if not compte_obj:
            errors.append({"row": i, "message": f"Compte introuvable: '{code_compte}'"}); continue

        errs_before = len(errors)
        montant = _dec(row.get("montant"), "montant", errors, i, min_val=Decimal("0.01"))
        if len(errors) > errs_before:
            continue

        date_recette = _parse_date(row.get("date_recette")) or datetime.date.today()

        raw_mois = str(row.get("mois", "")).strip().upper()
        mois = raw_mois if raw_mois in MOIS_VALUES else None

        valid.append({
            "date_recette": date_recette, "montant": montant, "libelle": libelle,
            "compte": compte_obj, "source": str(row.get("source", "")).strip(),
            "mois": mois,
        })
    return valid, errors


def _validate_depenses(rows, residence):
    comptes     = {c.code: c for c in CompteComptable.objects.filter(residence=residence)}
    categories  = {c.nom.strip().lower(): c for c in CategorieDepense.objects.filter(residence=residence)}
    fournisseurs = {f.nom.strip().lower(): f for f in Fournisseur.objects.filter(residence=residence)}
    valid, errors = [], []
    for i, row in enumerate(rows, start=2):
        libelle = str(row.get("libelle", "")).strip()
        if not libelle:
            errors.append({"row": i, "message": "libelle est obligatoire"}); continue

        code_compte = str(row.get("code_compte", "")).strip()
        if not code_compte:
            errors.append({"row": i, "message": "code_compte est obligatoire"}); continue
        compte_obj = comptes.get(code_compte)
        if not compte_obj:
            errors.append({"row": i, "message": f"Compte introuvable: '{code_compte}'"}); continue

        errs_before = len(errors)
        montant = _dec(row.get("montant"), "montant", errors, i, min_val=Decimal("0.01"))
        if len(errors) > errs_before:
            continue

        # Optional FKs
        categorie_obj = None
        raw_cat = str(row.get("categorie", "")).strip()
        if raw_cat:
            categorie_obj = categories.get(raw_cat.lower())
            if not categorie_obj:
                errors.append({"row": i, "message": f"Catégorie introuvable: '{raw_cat}'"}); continue

        fournisseur_obj = None
        raw_four = str(row.get("fournisseur", "")).strip()
        if raw_four:
            fournisseur_obj = fournisseurs.get(raw_four.lower())
            if not fournisseur_obj:
                errors.append({"row": i, "message": f"Fournisseur introuvable: '{raw_four}'"}); continue

        date_depense = _parse_date(row.get("date_depense")) or datetime.date.today()

        raw_mois = str(row.get("mois", "")).strip().upper()
        mois = raw_mois if raw_mois in MOIS_VALUES else None

        valid.append({
            "date_depense": date_depense, "montant": montant, "libelle": libelle,
            "compte": compte_obj, "categorie": categorie_obj, "fournisseur": fournisseur_obj,
            "facture_reference": str(row.get("facture_reference", "")).strip(),
            "mois": mois,
        })
    return valid, errors


# ── Inserters ─────────────────────────────────────────────────────────────────

def _insert_lots(valid_rows, residence):
    for d in valid_rows:
        Lot.objects.create(residence=residence, numero_lot=d["numero_lot"], type_lot=d["type_lot"],
                           etage_lot=d["etage_lot"], surface_lot=d["surface_lot"],
                           montant_ref=d["montant_ref"], groupe=d["groupe"], proprietaire=d["proprietaire"])
    return len(valid_rows)


def _insert_personnes(valid_rows, residence):
    for d in valid_rows:
        Personne.objects.get_or_create(residence=residence, nom=d["nom"], prenom=d["prenom"],
                                       defaults={"telephone": d["telephone"], "email": d["email"],
                                                 "type_personne": d["type_personne"]})
    return len(valid_rows)


def _insert_plan_comptable(valid_rows, residence):
    for d in valid_rows:
        CompteComptable.objects.create(residence=residence, code=d["code"],
                                       libelle=d["libelle"], type_compte=d["type_compte"])
    return len(valid_rows)


def _insert_appel_charge(valid_rows, residence):
    """AppelCharge.save() calls full_clean() which auto-generates code_fond."""
    for d in valid_rows:
        AppelCharge.objects.create(
            residence=residence,
            type_charge=d["type_charge"], exercice=d["exercice"], periode=d["periode"],
            nom_fond=d["nom_fond"], description_appel=d["description_appel"],
            date_emission=d["date_emission"],
        )
    return len(valid_rows)


def _insert_detail_appel(valid_rows, residence):
    """DetailAppelCharge.save() calls recalcule_statut() + full_clean()."""
    for d in valid_rows:
        DetailAppelCharge.objects.create(appel=d["appel"], lot=d["lot"], montant=d["montant"])
    return len(valid_rows)


def _insert_paiements(valid_rows, residence):
    """Paiement.save() automatically creates CaisseMouvement (DEBIT/Entrée)."""
    for d in valid_rows:
        Paiement.objects.create(
            residence=residence, lot=d["lot"],
            date_paiement=d["date_paiement"], montant=d["montant"], reference=d["reference"],
            mois=d["mois"], mode_paiement=d["mode_paiement"],
        )
    return len(valid_rows)


def _insert_affectation_paiement(valid_rows, residence):
    """After inserting affectations, rebuild montant_recu on each affected detail."""
    affected_details = set()
    for d in valid_rows:
        AffectationPaiement.objects.create(
            paiement=d["paiement"], detail=d["detail"], montant_affecte=d["montant_affecte"]
        )
        affected_details.add(d["detail"].id)
    # Rebuild montant_recu / statut on affected details
    for detail in DetailAppelCharge.objects.filter(id__in=affected_details):
        detail._rebuild_from_affectations()
    return len(valid_rows)


def _insert_recettes(valid_rows, residence):
    """Recette.save() automatically creates CaisseMouvement (DEBIT/Entrée)."""
    for d in valid_rows:
        Recette.objects.create(
            residence=residence, compte=d["compte"],
            date_recette=d["date_recette"], montant=d["montant"],
            libelle=d["libelle"], source=d["source"], mois=d["mois"],
        )
    return len(valid_rows)


def _insert_depenses(valid_rows, residence):
    """Depense.save() automatically creates CaisseMouvement (CREDIT/Sortie)."""
    for d in valid_rows:
        Depense.objects.create(
            residence=residence, compte=d["compte"],
            date_depense=d["date_depense"], montant=d["montant"], libelle=d["libelle"],
            categorie=d["categorie"], fournisseur=d["fournisseur"],
            facture_reference=d["facture_reference"], mois=d["mois"],
        )
    return len(valid_rows)


VALIDATORS = {
    "lots":                 _validate_lots,
    "personnes":            _validate_personnes,
    "plan-comptable":       _validate_plan_comptable,
    "appel-charge":         _validate_appel_charge,
    "detail-appel":         _validate_detail_appel,
    "paiements":            _validate_paiements,
    "affectation-paiement": _validate_affectation_paiement,
    "recettes":             _validate_recettes,
    "depenses":             _validate_depenses,
}
INSERTERS = {
    "lots":                 _insert_lots,
    "personnes":            _insert_personnes,
    "plan-comptable":       _insert_plan_comptable,
    "appel-charge":         _insert_appel_charge,
    "detail-appel":         _insert_detail_appel,
    "paiements":            _insert_paiements,
    "affectation-paiement": _insert_affectation_paiement,
    "recettes":             _insert_recettes,
    "depenses":             _insert_depenses,
}


# ── xlsx reader ───────────────────────────────────────────────────────────────

def _read_xlsx(file_obj):
    try:
        wb = openpyxl.load_workbook(file_obj, data_only=True)
    except Exception:
        raise ValueError("Fichier Excel invalide ou corrompu.")
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Le fichier est vide.")
    headers = [str(h).strip().lower() if h is not None else "" for h in rows[0]]
    data_rows = []
    for row in rows[1:]:
        if all(v is None or str(v).strip() == "" for v in row):
            continue
        data_rows.append({headers[j]: row[j] for j in range(len(headers))})
    return headers, data_rows


# ── Views ─────────────────────────────────────────────────────────────────────

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def import_template(request):
    dataset = request.GET.get("dataset", "lots")
    if dataset not in SPECS:
        return Response({"detail": f"Dataset inconnu: {dataset}"}, status=400)
    buf = _build_template(dataset)
    response = HttpResponse(
        buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="modele_{dataset}.xlsx"'
    return response


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def import_excel(request):
    residence = get_user_residence(request)
    if not residence:
        return Response({"detail": "Aucune résidence associée."}, status=400)

    dataset = request.POST.get("dataset", "").strip()
    action  = request.POST.get("action", "validate").strip()
    file    = request.FILES.get("file")

    if dataset not in SPECS:
        return Response({"detail": f"Dataset invalide. Choisissez parmi: {', '.join(SPECS)}"}, status=400)
    if not file:
        return Response({"detail": "Aucun fichier fourni."}, status=400)
    if action not in ("validate", "import"):
        return Response({"detail": "Action invalide (validate ou import)."}, status=400)

    try:
        headers, rows = _read_xlsx(file)
    except ValueError as e:
        return Response({"detail": str(e)}, status=400)

    spec = SPECS[dataset]
    missing_cols = [c for c in spec["required"] if c not in headers]
    if missing_cols:
        return Response({
            "detail": f"Colonnes manquantes: {', '.join(missing_cols)}. "
                      f"Colonnes attendues: {', '.join(spec['columns'])}"
        }, status=400)

    try:
        valid_rows, errors = VALIDATORS[dataset](rows, residence)
    except Exception as e:
        return Response({"detail": f"Erreur lors de la validation: {str(e)}"}, status=500)

    result = {
        "rows_detected": len(rows),
        "valid_rows":    len(valid_rows),
        "invalid_rows":  len(errors),
        "errors":        errors,
        "imported":      0,
    }

    if action == "import" and valid_rows:
        try:
            result["imported"] = INSERTERS[dataset](valid_rows, residence)
        except Exception as e:
            return Response({"detail": f"Erreur lors de l'insertion: {str(e)}"}, status=500)

    return Response(result)
