"""
Rapport Financier — vues séparées (JSON, Excel, PDF)
"""
import io
import datetime
import re
from decimal import Decimal

from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from django.http import HttpResponse
from django.db.models import Sum, OuterRef, Subquery

from django.db.models import Prefetch
from .models import CaisseMouvement, Depense, Recette, Paiement, Lot, DetailAppelCharge, AffectationPaiement, ArchivePaiement, ArchiveAffectationPaiement
from .views import get_user_residence


# ============================================================
# Helper partagé
# ============================================================
def _rapport_data(residence, date_debut, date_fin):
    mv_qs = CaisseMouvement.objects.filter(residence=residence)
    if date_debut:
        mv_qs = mv_qs.filter(date_mouvement__gte=date_debut)
    if date_fin:
        mv_qs = mv_qs.filter(date_mouvement__lte=date_fin)

    dep_qs = Depense.objects.select_related("compte", "categorie").filter(residence=residence)
    if date_debut:
        dep_qs = dep_qs.filter(date_depense__gte=date_debut)
    if date_fin:
        dep_qs = dep_qs.filter(date_depense__lte=date_fin)

    rec_qs = Recette.objects.select_related("compte").filter(residence=residence)
    if date_debut:
        rec_qs = rec_qs.filter(date_recette__gte=date_debut)
    if date_fin:
        rec_qs = rec_qs.filter(date_recette__lte=date_fin)

    pai_qs = Paiement.objects.filter(residence=residence)
    if date_debut:
        pai_qs = pai_qs.filter(date_paiement__gte=date_debut)
    if date_fin:
        pai_qs = pai_qs.filter(date_paiement__lte=date_fin)

    entrees        = mv_qs.filter(sens="DEBIT").aggregate(t=Sum("montant"))["t"]  or Decimal("0")
    sorties        = mv_qs.filter(sens="CREDIT").aggregate(t=Sum("montant"))["t"] or Decimal("0")
    total_paiements = pai_qs.aggregate(t=Sum("montant"))["t"] or Decimal("0")
    total_recettes  = rec_qs.aggregate(t=Sum("montant"))["t"] or Decimal("0")
    total_depenses  = dep_qs.aggregate(t=Sum("montant"))["t"] or Decimal("0")

    mouvements = list(mv_qs.order_by("date_mouvement").values(
        "id", "date_mouvement", "type_mouvement", "libelle", "montant", "sens", "commentaire",
    ))

    dep_par_compte    = {}
    dep_par_categorie = {}
    for d in dep_qs:
        code = f"{d.compte.code} — {d.compte.libelle}" if d.compte else "—"
        dep_par_compte[code]    = dep_par_compte.get(code, Decimal("0"))    + d.montant
        cat = d.categorie.nom if d.categorie else "Sans catégorie"
        dep_par_categorie[cat] = dep_par_categorie.get(cat, Decimal("0")) + d.montant

    rec_par_compte = {}
    for r in rec_qs:
        code = f"{r.compte.code} — {r.compte.libelle}" if r.compte else "—"
        rec_par_compte[code] = rec_par_compte.get(code, Decimal("0")) + r.montant

    # ── Situation par lot ───────────────────────────────────
    # Exclure les appels archivés (archive_comptable not null) pour ne pas
    # compter des montants déjà soldés. Utiliser montant_recu (persisté après
    # archivage des paiements) plutôt que Paiement.montant (supprimé).
    du_sub = (
        DetailAppelCharge.objects
        .filter(lot=OuterRef("pk"), appel__archive_comptable__isnull=True, archived=False)
        .values("lot_id")
        .annotate(s=Sum("montant"))
        .values("s")
    )
    paye_sub = (
        DetailAppelCharge.objects
        .filter(lot=OuterRef("pk"), appel__archive_comptable__isnull=True, archived=False)
        .values("lot_id")
        .annotate(s=Sum("montant_recu"))
        .values("s")
    )
    lots = Lot.objects.filter(residence=residence).select_related(
        "representant", "groupe"
    ).annotate(
        total_du   = Subquery(du_sub),
        total_paye = Subquery(paye_sub),
    ).order_by("groupe__nom_groupe", "numero_lot")

    situation_lots = []
    for lot in lots:
        du   = lot.total_du   or Decimal("0")
        paye = lot.total_paye or Decimal("0")
        reste = du - paye
        if reste < 0:
            statut = "TROP_PAYE"
        elif reste == 0 and paye > 0:
            statut = "A_JOUR"
        elif paye == 0 and du > 0:
            statut = "IMPAYE"
        elif reste > 0 and paye > 0:
            statut = "PARTIEL"
        else:
            statut = "A_JOUR"
        rep = lot.representant
        proprietaire = f"{rep.nom} {rep.prenom or ''}".strip() if rep else "—"
        telephone = rep.telephone if rep else ""
        email     = rep.email     if rep else ""
        situation_lots.append({
            "lot_id":       lot.id,
            "lot":          lot.numero_lot,
            "proprietaire": proprietaire,
            "telephone":    telephone,
            "email":        email,
            "total_du":     str(du),
            "total_paye":   str(paye),
            "reste":        str(reste),
            "statut":       statut,
        })

    return {
        "entrees":           entrees,
        "sorties":           sorties,
        "balance":           entrees - sorties,
        "total_paiements":   total_paiements,
        "total_recettes":    total_recettes,
        "total_depenses":    total_depenses,
        "mouvements":        mouvements,
        "dep_par_compte":    dep_par_compte,
        "dep_par_categorie": dep_par_categorie,
        "rec_par_compte":    rec_par_compte,
        "situation_lots":    situation_lots,
    }


# ============================================================
# JSON — données du rapport
# ============================================================
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def rapport_financier_json(request):
    residence = get_user_residence(request)
    if not residence:
        return Response({"detail": "Aucune résidence."}, status=400)

    date_debut = request.query_params.get("date_debut") or None
    date_fin   = request.query_params.get("date_fin")   or None
    d          = _rapport_data(residence, date_debut, date_fin)

    mouvements_out = []
    for m in d["mouvements"]:
        mouvements_out.append({
            "id":             m["id"],
            "date":           str(m["date_mouvement"]),
            "type_mouvement": m["type_mouvement"],
            "libelle":        m["libelle"],
            "montant":        str(m["montant"]),
            "sens":           m["sens"],
            "debit":          str(m["montant"]) if m["sens"] == "DEBIT"  else "0.00",
            "credit":         str(m["montant"]) if m["sens"] == "CREDIT" else "0.00",
            "commentaire":    m.get("commentaire") or "",
        })

    return Response({
        "residence":         residence.nom_residence,
        "entrees":           str(d["entrees"]),
        "sorties":           str(d["sorties"]),
        "balance":           str(d["balance"]),
        "total_paiements":   str(d["total_paiements"]),
        "total_recettes":    str(d["total_recettes"]),
        "total_depenses":    str(d["total_depenses"]),
        "mouvements":        mouvements_out,
        "dep_par_compte":    {k: str(v) for k, v in d["dep_par_compte"].items()},
        "dep_par_categorie": {k: str(v) for k, v in d["dep_par_categorie"].items()},
        "rec_par_compte":    {k: str(v) for k, v in d["rec_par_compte"].items()},
        "situation_lots":    d["situation_lots"],
    })


# ============================================================
# Export Excel
# ============================================================
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def rapport_financier_excel(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    residence = get_user_residence(request)
    if not residence:
        return Response({"detail": "Aucune résidence."}, status=400)

    date_debut = request.query_params.get("date_debut") or None
    date_fin   = request.query_params.get("date_fin")   or None
    d          = _rapport_data(residence, date_debut, date_fin)

    wb  = Workbook()
    thin   = Side(style="thin", color="E2E8F0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hdr_style(cell):
        cell.fill  = PatternFill("solid", fgColor="0F172A")
        cell.font  = Font(bold=True, color="FFFFFF", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    def data_cell(ws, row, col, value, fmt=None, bold=False):
        c = ws.cell(row=row, column=col, value=value)
        c.border = border
        if fmt:  c.number_format = fmt
        if bold: c.font = Font(bold=True)
        return c

    periode = f"{date_debut or '—'} → {date_fin or '—'}"

    # ── Feuille 1 : Résumé ──────────────────────────────────
    ws1 = wb.active
    ws1.title = "Résumé"
    ws1.column_dimensions["A"].width = 34
    ws1.column_dimensions["B"].width = 22
    ws1["A1"] = f"RAPPORT FINANCIER — {residence.nom_residence.upper()}"
    ws1["A1"].font = Font(bold=True, size=14)
    ws1["A2"] = f"Période : {periode}"
    ws1["A2"].font = Font(italic=True, color="64748B", size=10)
    ws1.append([])
    for col, h in enumerate(["Indicateur", "Montant (MAD)"], 1):
        hdr_style(ws1.cell(4, col, h))
    rows = [
        ("Total entrées",              float(d["entrees"])),
        ("Total sorties",              float(d["sorties"])),
        ("Balance de la période",      float(d["balance"])),
        ("Paiements copropriétaires",  float(d["total_paiements"])),
        ("Recettes externes",          float(d["total_recettes"])),
        ("Dépenses",                   float(d["total_depenses"])),
    ]
    for i, (label, val) in enumerate(rows, 5):
        data_cell(ws1, i, 1, label)
        data_cell(ws1, i, 2, val, fmt="#,##0.00")

    # ── Feuille 2 : Mouvements ──────────────────────────────
    ws2 = wb.create_sheet("Mouvements")
    for col, (h, w) in enumerate(zip(
        ["Date", "Type", "Libellé", "Débit (MAD)", "Crédit (MAD)", "Commentaire"],
        [14,     20,     38,        16,            16,             30]
    ), 1):
        ws2.column_dimensions[chr(64+col)].width = w
        hdr_style(ws2.cell(1, col, h))
    for i, m in enumerate(d["mouvements"], 2):
        debit  = float(m["montant"]) if m["sens"] == "DEBIT"  else None
        credit = float(m["montant"]) if m["sens"] == "CREDIT" else None
        data_cell(ws2, i, 1, str(m["date_mouvement"]))
        data_cell(ws2, i, 2, m["type_mouvement"])
        data_cell(ws2, i, 3, m["libelle"])
        data_cell(ws2, i, 4, debit,  fmt="#,##0.00")
        data_cell(ws2, i, 5, credit, fmt="#,##0.00")
        data_cell(ws2, i, 6, m.get("commentaire") or "")

    # ── Feuille 3 : Dépenses ────────────────────────────────
    ws3 = wb.create_sheet("Dépenses")
    ws3.column_dimensions["A"].width = 38
    ws3.column_dimensions["B"].width = 20
    ws3.column_dimensions["D"].width = 30
    ws3.column_dimensions["E"].width = 20
    for col, h in enumerate(["Compte comptable", "Montant (MAD)"], 1):
        hdr_style(ws3.cell(1, col, h))
    for i, (k, v) in enumerate(sorted(d["dep_par_compte"].items(), key=lambda x: -x[1]), 2):
        data_cell(ws3, i, 1, k)
        data_cell(ws3, i, 2, float(v), fmt="#,##0.00")
    hdr_style(ws3.cell(1, 4, "Catégorie"))
    hdr_style(ws3.cell(1, 5, "Montant (MAD)"))
    for i, (k, v) in enumerate(sorted(d["dep_par_categorie"].items(), key=lambda x: -x[1]), 2):
        data_cell(ws3, i, 4, k)
        data_cell(ws3, i, 5, float(v), fmt="#,##0.00")

    # ── Feuille 4 : Recettes ────────────────────────────────
    ws4 = wb.create_sheet("Recettes & Paiements")
    ws4.column_dimensions["A"].width = 38
    ws4.column_dimensions["B"].width = 20
    for col, h in enumerate(["Source", "Montant (MAD)"], 1):
        hdr_style(ws4.cell(1, col, h))
    data_cell(ws4, 2, 1, "Paiements copropriétaires", bold=True)
    data_cell(ws4, 2, 2, float(d["total_paiements"]), fmt="#,##0.00", bold=True)
    for i, (k, v) in enumerate(sorted(d["rec_par_compte"].items(), key=lambda x: -x[1]), 3):
        data_cell(ws4, i, 1, k)
        data_cell(ws4, i, 2, float(v), fmt="#,##0.00")

    # ── Feuille 5 : Situation des lots ──────────────────────
    ws5 = wb.create_sheet("Situation des lots")
    for col, (h, w) in enumerate(zip(
        ["Lot", "Propriétaire", "Total dû (MAD)", "Total payé (MAD)", "Reste (MAD)", "Statut"],
        [12,    28,              18,               18,                 16,            14]
    ), 1):
        ws5.column_dimensions[chr(64+col)].width = w
        hdr_style(ws5.cell(1, col, h))
    for i, lot in enumerate(d["situation_lots"], 2):
        reste = float(lot["reste"])
        data_cell(ws5, i, 1, lot["lot"])
        data_cell(ws5, i, 2, lot["proprietaire"])
        data_cell(ws5, i, 3, float(lot["total_du"]),   fmt="#,##0.00")
        data_cell(ws5, i, 4, float(lot["total_paye"]), fmt="#,##0.00")
        c5 = data_cell(ws5, i, 5, reste, fmt="#,##0.00")
        if reste > 0:
            c5.font = Font(bold=True, color="DC2626")
        data_cell(ws5, i, 6, "À jour" if lot["statut"] == "A_JOUR" else "En retard")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    slug  = re.sub(r"[^a-z0-9]", "_", residence.nom_residence.lower())
    today = datetime.date.today().strftime("%Y%m%d")
    resp  = HttpResponse(
        buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="rapport_{slug}_{today}.xlsx"'
    return resp


# ============================================================
# Export PDF
# ============================================================
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def rapport_financier_pdf(request):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable,
    )
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER

    residence = get_user_residence(request)
    if not residence:
        return Response({"detail": "Aucune résidence."}, status=400)

    date_debut = request.query_params.get("date_debut") or None
    date_fin   = request.query_params.get("date_fin")   or None
    d          = _rapport_data(residence, date_debut, date_fin)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=1.8*cm, rightMargin=1.8*cm,
        topMargin=2*cm, bottomMargin=2*cm,
    )
    styles = getSampleStyleSheet()
    S = lambda name, **kw: ParagraphStyle(name, parent=styles["Normal"], **kw)
    title_s   = S("T",  fontSize=17, fontName="Helvetica-Bold", spaceAfter=3)
    sub_s     = S("S",  fontSize=10, textColor=colors.HexColor("#64748B"), spaceAfter=14)
    section_s = S("H",  fontSize=11, fontName="Helvetica-Bold", spaceAfter=6, spaceBefore=14)
    footer_s  = S("F",  fontSize=8,  textColor=colors.HexColor("#94A3B8"), alignment=TA_CENTER)

    DARK  = colors.HexColor("#0F172A")
    GOLD  = colors.HexColor("#C9A84C")
    LIGHT = colors.HexColor("#F8FAFC")
    GREEN = colors.HexColor("#059669")
    RED   = colors.HexColor("#DC2626")

    def money(v): return f"{float(v):,.2f} MAD"
    periode = f"{date_debut or '—'} → {date_fin or '—'}"

    BASE_TS = lambda: TableStyle([
        ("BACKGROUND",   (0,0), (-1,0), DARK),
        ("TEXTCOLOR",    (0,0), (-1,0), colors.white),
        ("FONTNAME",     (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE",     (0,0), (-1,-1), 9),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [LIGHT, colors.white]),
        ("GRID",         (0,0), (-1,-1), 0.4, colors.HexColor("#E2E8F0")),
        ("ALIGN",        (-1,0), (-1,-1), "RIGHT"),
        ("LEFTPADDING",  (0,0), (-1,-1), 7),
        ("RIGHTPADDING", (0,0), (-1,-1), 7),
        ("TOPPADDING",   (0,0), (-1,-1), 5),
        ("BOTTOMPADDING",(0,0), (-1,-1), 5),
    ])

    elems = []
    elems.append(Paragraph("Rapport Financier", title_s))
    elems.append(Paragraph(f"{residence.nom_residence} · Période : {periode}", sub_s))
    elems.append(HRFlowable(width="100%", thickness=2, color=GOLD, spaceAfter=10))

    # Résumé
    elems.append(Paragraph("Résumé financier", section_s))
    ts = BASE_TS()
    ts.add("FONTNAME",  (0,3), (-1,3), "Helvetica-Bold")
    ts.add("TEXTCOLOR", (1,3), (1,3), GREEN if d["balance"] >= 0 else RED)
    t = Table([
        ["Indicateur", "Montant (MAD)"],
        ["Total entrées",             money(d["entrees"])],
        ["Total sorties",             money(d["sorties"])],
        ["Balance de la période",     money(d["balance"])],
        ["Paiements copropriétaires", money(d["total_paiements"])],
        ["Recettes externes",         money(d["total_recettes"])],
        ["Dépenses",                  money(d["total_depenses"])],
    ], colWidths=[11*cm, 6*cm])
    t.setStyle(ts)
    elems.append(t)

    # Mouvements (max 60)
    elems.append(Paragraph("Mouvements de caisse", section_s))
    mv_rows = [["Date", "Type", "Libellé", "Débit", "Crédit"]]
    for m in d["mouvements"][:60]:
        mv_rows.append([
            str(m["date_mouvement"]),
            m["type_mouvement"],
            (m["libelle"] or "")[:38],
            money(m["montant"]) if m["sens"] == "DEBIT"  else "",
            money(m["montant"]) if m["sens"] == "CREDIT" else "",
        ])
    t2 = Table(mv_rows, colWidths=[2.4*cm, 3*cm, 6.4*cm, 2.8*cm, 2.8*cm])
    t2.setStyle(BASE_TS())
    elems.append(t2)
    if len(d["mouvements"]) > 60:
        elems.append(Paragraph(
            f"… {len(d['mouvements'])-60} ligne(s) supplémentaire(s) dans l'export Excel.", sub_s))

    # Dépenses par compte
    if d["dep_par_compte"]:
        elems.append(Paragraph("Dépenses par compte", section_s))
        rows = [["Compte comptable", "Montant (MAD)"]] + [
            [k, money(v)] for k, v in sorted(d["dep_par_compte"].items(), key=lambda x: -x[1])
        ]
        t3 = Table(rows, colWidths=[11.5*cm, 6*cm])
        t3.setStyle(BASE_TS())
        elems.append(t3)

    # Recettes
    elems.append(Paragraph("Recettes & paiements", section_s))
    rows4 = [["Source", "Montant (MAD)"],
             ["Paiements copropriétaires", money(d["total_paiements"])]]
    rows4 += [[k, money(v)] for k, v in sorted(d["rec_par_compte"].items(), key=lambda x: -x[1])]
    t4 = Table(rows4, colWidths=[11.5*cm, 6*cm])
    t4.setStyle(BASE_TS())
    elems.append(t4)

    # Situation des lots
    if d["situation_lots"]:
        elems.append(Paragraph("Situation des lots", section_s))
        lots_rows = [["Lot", "Propriétaire", "Total dû", "Total payé", "Reste", "Statut"]]
        for lot in d["situation_lots"]:
            lots_rows.append([
                lot["lot"],
                lot["proprietaire"],
                money(lot["total_du"]),
                money(lot["total_paye"]),
                money(lot["reste"]),
                "À jour" if lot["statut"] == "A_JOUR" else "En retard",
            ])
        t5 = Table(lots_rows, colWidths=[1.5*cm, 4.5*cm, 3*cm, 3*cm, 3*cm, 2.4*cm])
        ts5 = BASE_TS()
        # Color "En retard" rows in reste column
        for row_idx, lot in enumerate(d["situation_lots"], 1):
            if float(lot["reste"]) > 0:
                ts5.add("TEXTCOLOR", (4, row_idx), (4, row_idx), RED)
                ts5.add("FONTNAME",  (4, row_idx), (4, row_idx), "Helvetica-Bold")
        t5.setStyle(ts5)
        elems.append(t5)

    elems.append(Spacer(1, 18))
    elems.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#E2E8F0")))
    elems.append(Paragraph(
        f"Généré le {datetime.date.today().strftime('%d/%m/%Y')} · Syndic Pro", footer_s))

    doc.build(elems)
    buf.seek(0)

    slug  = re.sub(r"[^a-z0-9]", "_", residence.nom_residence.lower())
    today = datetime.date.today().strftime("%Y%m%d")
    resp  = HttpResponse(buf.read(), content_type="application/pdf")
    resp["Content-Disposition"] = f'attachment; filename="rapport_{slug}_{today}.pdf"'
    return resp


# ============================================================
# Export Situation des paiements — Excel
# ============================================================
STATUT_LABELS = {
    "A_JOUR":    "À jour",
    "PARTIEL":   "Partiel",
    "IMPAYE":    "Impayé",
    "TROP_PAYE": "Trop payé",
}

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def situation_export_excel(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    residence = get_user_residence(request)
    if not residence:
        return Response({"detail": "Aucune résidence."}, status=400)

    d    = _rapport_data(residence, None, None)
    lots = d["situation_lots"]

    wb  = Workbook()
    ws  = wb.active
    ws.title = "Situation paiements"
    thin   = Side(style="thin", color="E2E8F0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    cols = [
        ("Lot",           10), ("Propriétaire", 26), ("Téléphone",   16),
        ("Total dû",      16), ("Total payé",   16), ("Reste",       16),
        ("Statut",        14),
    ]
    for ci, (h, w) in enumerate(cols, 1):
        ws.column_dimensions[chr(64 + ci)].width = w
        c = ws.cell(1, ci, h)
        c.fill      = PatternFill("solid", fgColor="0F172A")
        c.font      = Font(bold=True, color="FFFFFF", size=10)
        c.alignment = Alignment(horizontal="center")
        c.border    = border

    RED   = "DC2626"
    GREEN = "059669"
    AMBER = "D97706"
    BLUE  = "1D4ED8"

    total_du = total_paye = total_reste = Decimal("0")
    for ri, lot in enumerate(lots, 2):
        reste = Decimal(lot["reste"])
        du    = Decimal(lot["total_du"])
        paye  = Decimal(lot["total_paye"])
        total_du   += du
        total_paye += paye
        total_reste += reste

        color_reste = RED if reste > 0 else (BLUE if reste < 0 else GREEN)
        statut_lbl  = STATUT_LABELS.get(lot["statut"], lot["statut"])

        for ci, val in enumerate([
            lot["lot"], lot["proprietaire"], lot["telephone"],
            float(du), float(paye), float(reste), statut_lbl,
        ], 1):
            c = ws.cell(ri, ci, val)
            c.border = border
            if ci in (4, 5, 6) and val is not None:
                c.number_format = "#,##0.00"
            if ci == 6:
                c.font = Font(bold=True, color=color_reste)

    # Totaux
    last = len(lots) + 2
    ws.cell(last, 1, "TOTAL").font = Font(bold=True)
    for ci, val in [(4, float(total_du)), (5, float(total_paye)), (6, float(total_reste))]:
        c = ws.cell(last, ci, val)
        c.number_format = "#,##0.00"
        c.font = Font(bold=True, color=RED if ci == 6 and total_reste > 0 else GREEN)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    slug  = re.sub(r"[^a-z0-9]", "_", residence.nom_residence.lower())
    today = datetime.date.today().strftime("%Y%m%d")
    resp  = HttpResponse(buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = f'attachment; filename="situation_{slug}_{today}.xlsx"'
    return resp


# ============================================================
# Export Situation des paiements — PDF
# ============================================================
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def situation_export_pdf(request):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    residence = get_user_residence(request)
    if not residence:
        return Response({"detail": "Aucune résidence."}, status=400)

    d    = _rapport_data(residence, None, None)
    lots = d["situation_lots"]

    def money(v): return f"{float(v):,.2f}"

    DARK  = colors.HexColor("#0F172A")
    RED   = colors.HexColor("#DC2626")
    GREEN = colors.HexColor("#059669")
    AMBER = colors.HexColor("#D97706")
    BLUE  = colors.HexColor("#1D4ED8")
    LIGHT = colors.HexColor("#F8FAFC")

    styles = getSampleStyleSheet()
    title_s = ParagraphStyle("T", parent=styles["Normal"], fontSize=14, fontName="Helvetica-Bold", spaceAfter=4)
    sub_s   = ParagraphStyle("S", parent=styles["Normal"], fontSize=9,  textColor=colors.HexColor("#64748B"), spaceAfter=10)

    headers = ["Lot", "Propriétaire", "Téléphone", "Total dû", "Total payé", "Reste", "Statut"]
    rows    = [headers]
    total_du = total_paye = total_reste = Decimal("0")

    for lot in lots:
        reste = Decimal(lot["reste"])
        du    = Decimal(lot["total_du"])
        paye  = Decimal(lot["total_paye"])
        total_du   += du
        total_paye += paye
        total_reste += reste
        rows.append([
            lot["lot"], lot["proprietaire"], lot["telephone"] or "—",
            money(du), money(paye), money(reste),
            STATUT_LABELS.get(lot["statut"], lot["statut"]),
        ])

    rows.append(["TOTAL", "", "", money(total_du), money(total_paye), money(total_reste), ""])

    ts = TableStyle([
        ("BACKGROUND",    (0, 0), (-1, 0), DARK),
        ("TEXTCOLOR",     (0, 0), (-1, 0), colors.white),
        ("FONTNAME",      (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME",      (0,-1), (-1,-1), "Helvetica-Bold"),
        ("FONTSIZE",      (0, 0), (-1,-1), 8),
        ("ROWBACKGROUNDS",(0, 1), (-1,-2), [LIGHT, colors.white]),
        ("GRID",          (0, 0), (-1,-1), 0.4, colors.HexColor("#E2E8F0")),
        ("ALIGN",         (3, 0), (5,-1), "RIGHT"),
        ("LEFTPADDING",   (0, 0), (-1,-1), 5),
        ("RIGHTPADDING",  (0, 0), (-1,-1), 5),
        ("TOPPADDING",    (0, 0), (-1,-1), 4),
        ("BOTTOMPADDING", (0, 0), (-1,-1), 4),
    ])
    # Color reste column per row
    for ri, lot in enumerate(lots, 1):
        reste = float(lot["reste"])
        c = RED if reste > 0 else (BLUE if reste < 0 else GREEN)
        ts.add("TEXTCOLOR", (5, ri), (5, ri), c)
        if reste != 0:
            ts.add("FONTNAME", (5, ri), (5, ri), "Helvetica-Bold")

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
        leftMargin=1.5*cm, rightMargin=1.5*cm, topMargin=1.8*cm, bottomMargin=1.8*cm)

    taux = (float(total_paye) / float(total_du) * 100) if total_du else 0
    table = Table(rows, colWidths=[1.5*cm, 5*cm, 3.5*cm, 3*cm, 3*cm, 3*cm, 2.5*cm])
    table.setStyle(ts)

    elems = [
        Paragraph("Situation des paiements", title_s),
        Paragraph(
            f"{residence.nom_residence} · "
            f"Total dû : {money(total_du)} MAD · "
            f"Recouvré : {money(total_paye)} MAD · "
            f"Taux : {taux:.1f}%", sub_s),
        table,
    ]
    doc.build(elems)
    buf.seek(0)
    slug  = re.sub(r"[^a-z0-9]", "_", residence.nom_residence.lower())
    today = datetime.date.today().strftime("%Y%m%d")
    resp  = HttpResponse(buf.read(), content_type="application/pdf")
    resp["Content-Disposition"] = f'attachment; filename="situation_{slug}_{today}.pdf"'
    return resp


# ============================================================
# Situation paiements — analyse timeline avec report inter-années
# ============================================================
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def situation_paiements_view(request):
    """
    Logique cumulative avec report, séparée par type (CHARGE ou FOND).

    Le type d'un paiement est déduit de ses affectations :
    - Paiement avec affectations CHARGE → pool CHARGE
    - Paiement avec affectations FOND   → pool FOND
    - Paiement sans affectation (unventilé) → inclus dans le pool CHARGE seulement

    Exemple : charges CHARGE 250/an (2026-2027-2028), paiement 300 MAD (CHARGE)
      → 2026 : 250 couverts (barre pleine)
      → 2027 : 50 couverts  (report du surplus)
      → 2028 : 0
    """
    residence = get_user_residence(request)
    if not residence:
        return Response({"detail": "Aucune résidence."}, status=400)

    try:
        year = int(request.query_params.get("year", 0)) or datetime.date.today().year
    except (TypeError, ValueError):
        year = datetime.date.today().year

    type_charge = request.query_params.get("type_charge", "CHARGE")
    if type_charge not in ("CHARGE", "FOND"):
        type_charge = "CHARGE"

    # Filtre optionnel sur un appel de fond spécifique
    appel_id = request.query_params.get("appel_id")
    try:
        appel_id = int(appel_id) if appel_id else None
    except (TypeError, ValueError):
        appel_id = None

    cutoff = datetime.date(year, 12, 31)

    lots_qs = (
        Lot.objects
        .filter(residence=residence)
        .select_related("representant", "groupe")
        .order_by("groupe__nom_groupe", "numero_lot")
    )
    lot_ids = list(lots_qs.values_list("id", flat=True))

    # ── Charges par (lot_id, exercice) ───────────────────────────
    from django.db.models import Sum as DSum
    if appel_id:
        charges_qs = DetailAppelCharge.objects.filter(lot__in=lot_ids, appel__id=appel_id)
    else:
        charges_qs = DetailAppelCharge.objects.filter(lot__in=lot_ids, appel__type_charge=type_charge)
    charges_rows = charges_qs.values("lot_id", "appel__exercice").annotate(total=DSum("montant"))
    charges_map = {}  # {lot_id: {exercice: total}}
    for row in charges_rows:
        charges_map.setdefault(row["lot_id"], {})[row["appel__exercice"]] = float(row["total"])

    # ── Montants ventilés live ────────────────────────────────────
    aff_map = {}  # {lot_id: [(date, pmt_id, montant)]}
    if appel_id:
        aff_filter = dict(paiement__lot__in=lot_ids, detail__appel__id=appel_id)
    else:
        aff_filter = dict(paiement__lot__in=lot_ids, detail__appel__type_charge=type_charge)
    for row in (
        AffectationPaiement.objects
        .filter(**aff_filter)
        .values("paiement__lot_id", "paiement__id", "paiement__date_paiement")
        .annotate(total=DSum("montant_affecte"))
        .order_by("paiement__lot_id", "paiement__date_paiement", "paiement__id")
    ):
        aff_map.setdefault(row["paiement__lot_id"], []).append((
            row["paiement__date_paiement"],
            row["paiement__id"],
            float(row["total"]),
        ))

    # ── Montants ventilés archivés ────────────────────────────────
    if appel_id:
        arch_aff_filter = dict(archive_paiement__lot__in=lot_ids, detail__appel__id=appel_id)
    else:
        arch_aff_filter = dict(archive_paiement__lot__in=lot_ids, detail__appel__type_charge=type_charge)
    for row in (
        ArchiveAffectationPaiement.objects
        .filter(**arch_aff_filter)
        .values("archive_paiement__lot_id", "archive_paiement__id", "archive_paiement__date_paiement")
        .annotate(total=DSum("montant_affecte"))
        .order_by("archive_paiement__lot_id", "archive_paiement__date_paiement", "archive_paiement__id")
    ):
        aff_map.setdefault(row["archive_paiement__lot_id"], []).append((
            row["archive_paiement__date_paiement"],
            -row["archive_paiement__id"],
            float(row["total"]),
        ))

    # ── Paiements NON ventilés live — inclus dans CHARGE seulement (pas si appel_id) ──
    unvent_map = {}  # {lot_id: [(date, pmt_id, montant)]}
    if type_charge == "CHARGE" and not appel_id:
        has_aff = set(
            AffectationPaiement.objects
            .filter(paiement__lot__in=lot_ids)
            .values_list("paiement_id", flat=True)
            .distinct()
        )
        for row in (
            Paiement.objects
            .filter(lot__in=lot_ids)
            .exclude(id__in=has_aff)
            .values("lot_id", "id", "date_paiement", "montant")
            .order_by("date_paiement", "id")
        ):
            unvent_map.setdefault(row["lot_id"], []).append((
                row["date_paiement"], row["id"], float(row["montant"]),
            ))

        # ── Paiements NON ventilés archivés ───────────────────────────────
        has_arch_aff = set(
            ArchiveAffectationPaiement.objects
            .filter(archive_paiement__lot__in=lot_ids)
            .values_list("archive_paiement_id", flat=True)
            .distinct()
        )
        for row in (
            ArchivePaiement.objects
            .filter(lot__in=lot_ids)
            .exclude(id__in=has_arch_aff)
            .values("lot_id", "id", "date_paiement", "montant")
            .order_by("date_paiement", "id")
        ):
            unvent_map.setdefault(row["lot_id"], []).append((
                row["date_paiement"], -row["id"], float(row["montant"]),
            ))

    # ── Fond paiements directs (pour grille FOND) ─────────────────
    # Inclus uniquement quand appel_id est fourni (mode mois-direct)
    _MOIS_CODES = ["JAN","FEV","MAR","AVR","MAI","JUN","JUL","AOU","SEP","OCT","NOV","DEC"]
    fond_pmt_map = {}   # {lot_id: [{id, mois, montant, date}]}
    if type_charge == "FOND" and appel_id:
        for row in (
            AffectationPaiement.objects
            .filter(paiement__lot__in=lot_ids, detail__appel__id=appel_id)
            .values(
                "paiement__lot_id", "paiement__id",
                "paiement__mois", "paiement__date_paiement",
                "montant_affecte",
            )
            .order_by("paiement__lot_id", "paiement__date_paiement")
        ):
            # Si mois non renseigné, le déduire de date_paiement
            mois = row["paiement__mois"]
            if not mois:
                dt = row["paiement__date_paiement"]
                mois = _MOIS_CODES[dt.month - 1] if dt else ""
            fond_pmt_map.setdefault(row["paiement__lot_id"], []).append({
                "id":      row["paiement__id"],
                "mois":    mois,
                "montant": float(row["montant_affecte"]),
                "date":    str(row["paiement__date_paiement"]),
            })

    # ── Construction résultat ─────────────────────────────────
    result = []
    for lot in lots_qs:
        rep = lot.representant
        nom = f"{rep.nom} {rep.prenom or ''}".strip() if rep else "—"

        charges_by_year = charges_map.get(lot.id, {})
        total_du_year   = charges_by_year.get(year, 0.0)

        # Combiner paiements ventilés (type) + paiements unventilés (CHARGE seulement)
        raw_payments = sorted(
            aff_map.get(lot.id, []) + unvent_map.get(lot.id, []),
            key=lambda x: (x[0], x[1])
        )

        # Filtrer : uniquement les paiements jusqu'à la fin de l'année demandée
        payments_up_to = [(d, pid, amt) for d, pid, amt in raw_payments if d <= cutoff]

        # Cumul des charges AVANT l'année demandée
        cumul_before = sum(v for y, v in charges_by_year.items() if y < year)

        # Cumul payé (vers ce type) jusqu'à fin de l'année
        cumul_paid = sum(amt for _, _, amt in payments_up_to)

        available = cumul_paid - cumul_before
        effective = max(0.0, min(total_du_year, available))

        # Segments pour la timeline
        segments = []
        if effective > 0:
            to_skip   = cumul_before
            remaining = total_du_year

            for date, _, p_amt in payments_up_to:
                if remaining <= 0:
                    break
                if to_skip > 0:
                    if p_amt <= to_skip:
                        to_skip -= p_amt
                        continue
                    else:
                        p_amt  -= to_skip
                        to_skip = 0
                contribution = min(p_amt, remaining)
                if contribution > 0:
                    segments.append({
                        "montant": round(contribution, 2),
                        "date":    str(date),
                    })
                    remaining -= contribution

        lot_entry = {
            "lot_id":    lot.id,
            "lot":       lot.numero_lot,
            "nom":       nom,
            "groupe":    lot.groupe.nom_groupe if lot.groupe else "",
            "total_du":  total_du_year,
            "paiements": segments,
        }
        # Pour FOND avec appel_id : ajouter les paiements directs avec mois/id
        if type_charge == "FOND" and appel_id:
            fp = fond_pmt_map.get(lot.id, [])
            lot_entry["fond_paiements"] = fp
            lot_entry["total_paye"] = sum(p["montant"] for p in fp)
        result.append(lot_entry)

    # Années disponibles (tous les appels du type, pas uniquement l'appel sélectionné)
    available_years = sorted(
        DetailAppelCharge.objects
        .filter(lot__in=lot_ids, appel__type_charge=type_charge)
        .values_list("appel__exercice", flat=True)
        .distinct()
    )

    return Response({"years": available_years, "lots": result})


# ============================================================
# Années avec activité financière — GET /api/annees-activite/
# ============================================================
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def annees_activite(request):
    """Retourne toutes les années ayant au moins une dépense, recette ou paiement."""
    residence = get_user_residence(request)
    if not residence:
        return Response([])

    years = set()
    years.update(Depense.objects.filter(residence=residence)
                 .values_list("date_depense__year", flat=True).distinct())
    years.update(Recette.objects.filter(residence=residence)
                 .values_list("date_recette__year", flat=True).distinct())
    years.update(Paiement.objects.filter(residence=residence)
                 .values_list("date_paiement__year", flat=True).distinct())
    years.discard(None)
    return Response(sorted(years, reverse=True))


# ============================================================
# Envoi d'email — POST /api/send-email/
# ============================================================
@api_view(["POST"])
@permission_classes([IsAuthenticated])
def send_email_view(request):
    from django.core.mail import get_connection, EmailMessage
    from django.conf import settings as django_settings
    from .models import Residence

    to_email = request.data.get("to", "").strip()
    subject  = request.data.get("subject", "").strip()
    body     = request.data.get("body", "").strip()

    if not to_email or not subject or not body:
        return Response({"error": "Champs manquants (to, subject, body)."}, status=400)

    # Récupérer les SMTP de la résidence active de l'utilisateur
    from_email = None
    connection = None
    try:
        membership = request.user.memberships.select_related("residence").first()
        residence  = membership.residence if membership else None
        if residence and residence.email and residence.email_password:
            connection = get_connection(
                backend="django.core.mail.backends.smtp.EmailBackend",
                host="smtp.gmail.com",
                port=587,
                username=residence.email,
                password=residence.email_password,
                use_tls=True,
                fail_silently=False,
            )
            from_email = residence.email
    except Exception:
        pass  # fallback sur les settings Django

    try:
        msg = EmailMessage(
            subject=subject,
            body=body,
            from_email=from_email or django_settings.DEFAULT_FROM_EMAIL,
            to=[to_email],
            connection=connection,
        )
        msg.send(fail_silently=False)
        return Response({"ok": True})
    except Exception as e:
        return Response({"error": str(e)}, status=500)


# ============================================================
# Helpers partagés pour l'export Saisie Grille
# ============================================================
MOIS_FR = ["Janvier","Février","Mars","Avril","Mai","Juin",
           "Juillet","Août","Septembre","Octobre","Novembre","Décembre"]
MOIS_SHORT = ["Jan","Fév","Mar","Avr","Mai","Jun","Jul","Aoû","Sep","Oct","Nov","Déc"]

def _saisie_grille_data(residence, year, month, type_charge, appel_id=None):
    """
    Retourne un dict avec:
      - lots: [{lot, nom, paid[12], total_du, total_paye, reste}]
      - total_paiements_mois, total_depenses_mois, balance
      - depenses: [{date, libelle, categorie, fournisseur, montant}]
    Pour FOND + appel_id : logique mois-direct (pas de mensualités).
    """
    MOIS_CODES = ["JAN","FEV","MAR","AVR","MAI","JUN","JUL","AOU","SEP","OCT","NOV","DEC"]
    import calendar as cal
    last_day = cal.monthrange(year, month)[1]
    cutoff   = datetime.date(year, month, last_day)
    deb_mois = datetime.date(year, month, 1)

    lots_qs = (
        Lot.objects
        .filter(residence=residence)
        .select_related("representant", "groupe")
        .order_by("groupe__nom_groupe", "numero_lot")
    )
    lot_ids = list(lots_qs.values_list("id", flat=True))

    from django.db.models import Sum as DSum

    # Charges par lot
    if appel_id:
        charges_qs = DetailAppelCharge.objects.filter(lot__in=lot_ids, appel__id=appel_id)
    else:
        charges_qs = DetailAppelCharge.objects.filter(
            lot__in=lot_ids, appel__type_charge=type_charge, appel__exercice=year)
    charges_map = {r["lot_id"]: float(r["total"])
                   for r in charges_qs.values("lot_id").annotate(total=DSum("montant"))}

    # KPI mois : paiements encaissés dans le mois (tous types)
    total_paiements_mois = float(
        Paiement.objects
        .filter(residence=residence, date_paiement__gte=deb_mois, date_paiement__lte=cutoff)
        .aggregate(t=DSum("montant"))["t"] or 0
    )

    # Dépenses du mois
    dep_qs = (
        Depense.objects
        .select_related("categorie", "fournisseur")
        .filter(residence=residence, date_depense__gte=deb_mois, date_depense__lte=cutoff)
        .order_by("date_depense")
    )
    depenses = []
    total_depenses_mois = 0.0
    for d in dep_qs:
        m = float(d.montant)
        total_depenses_mois += m
        fournisseur = ""
        if d.fournisseur:
            f = d.fournisseur
            fournisseur = f.nom_societe or f"{f.nom} {f.prenom or ''}".strip()
        depenses.append({
            "date":        str(d.date_depense),
            "libelle":     d.libelle,
            "categorie":   d.categorie.nom if d.categorie else "",
            "fournisseur": fournisseur,
            "montant":     m,
        })

    lots_out = []

    if type_charge == "FOND" and appel_id:
        # ── FOND mode : logique mois-direct ──────────────────────────
        fond_pmts = (
            AffectationPaiement.objects
            .filter(paiement__lot__in=lot_ids, detail__appel__id=appel_id)
            .values(
                "paiement__lot_id", "paiement__id",
                "paiement__mois", "paiement__date_paiement",
                "montant_affecte",
            )
            .order_by("paiement__lot_id", "paiement__date_paiement")
        )
        fond_pmt_map = {}   # {lot_id: [{id, mois, montant, date}]}
        for row in fond_pmts:
            # Si mois non renseigné, le déduire de date_paiement
            mois = row["paiement__mois"]
            if not mois:
                dt = row["paiement__date_paiement"]
                mois = MOIS_CODES[dt.month - 1] if dt else ""
            fond_pmt_map.setdefault(row["paiement__lot_id"], []).append({
                "id":     row["paiement__id"],
                "mois":   mois,
                "montant": float(row["montant_affecte"]),
                "date":   row["paiement__date_paiement"],
            })

        for lot in lots_qs:
            rep = lot.representant
            nom = f"{rep.nom} {rep.prenom or ''}".strip() if rep else "—"
            total_du   = charges_map.get(lot.id, 0.0)
            pmts       = fond_pmt_map.get(lot.id, [])
            total_paye = sum(p["montant"] for p in pmts)
            reste      = total_du - total_paye

            paid = [any(p["mois"] == MOIS_CODES[i] for p in pmts) for i in range(12)]
            paid_before = [
                any(p["mois"] == MOIS_CODES[i] and p["date"] < deb_mois for p in pmts)
                for i in range(12)
            ]
            paid_this_month = [
                any(p["mois"] == MOIS_CODES[i] and deb_mois <= p["date"] <= cutoff for p in pmts)
                for i in range(12)
            ]
            lots_out.append({
                "lot": lot.numero_lot, "nom": nom,
                "paid": paid, "paid_before": paid_before, "paid_this_month": paid_this_month,
                "total_du": total_du, "total_paye": total_paye, "reste": reste,
                "fond_paiements": [
                    {"id": p["id"], "mois": p["mois"],
                     "montant": p["montant"], "date": str(p["date"])}
                    for p in pmts
                ],
            })

    else:
        # ── CHARGE mode : logique carry-over (mensualités) ────────────
        aff_rows = (
            AffectationPaiement.objects
            .filter(
                paiement__lot__in=lot_ids,
                paiement__date_paiement__lte=cutoff,
                detail__appel__type_charge=type_charge,
                detail__appel__exercice=year,
            )
            .values("paiement__lot_id")
            .annotate(total=DSum("montant_affecte"))
        )
        paye_map = {r["paiement__lot_id"]: float(r["total"]) for r in aff_rows}

        aff_before_rows = (
            AffectationPaiement.objects
            .filter(
                paiement__lot__in=lot_ids,
                paiement__date_paiement__lt=deb_mois,
                detail__appel__type_charge=type_charge,
                detail__appel__exercice=year,
            )
            .values("paiement__lot_id")
            .annotate(total=DSum("montant_affecte"))
        )
        paye_before_map = {r["paiement__lot_id"]: float(r["total"]) for r in aff_before_rows}

        for lot in lots_qs:
            rep = lot.representant
            nom = f"{rep.nom} {rep.prenom or ''}".strip() if rep else "—"
            total_du    = charges_map.get(lot.id, 0.0)
            total_paye  = paye_map.get(lot.id, 0.0)
            reste       = total_du - total_paye
            months_covered = (total_paye / total_du * 12) if total_du > 0 else 0.0
            paye_before    = paye_before_map.get(lot.id, 0.0)
            months_before  = (paye_before / total_du * 12) if total_du > 0 else 0.0
            paid            = [i < months_covered for i in range(12)]
            paid_before     = [i < months_before for i in range(12)]
            paid_this_month = [(months_before <= i < months_covered) for i in range(12)]
            lots_out.append({
                "lot": lot.numero_lot, "nom": nom,
                "paid": paid, "paid_before": paid_before, "paid_this_month": paid_this_month,
                "total_du": total_du, "total_paye": total_paye, "reste": reste,
            })

    return {
        "lots": lots_out,
        "total_paiements_mois": total_paiements_mois,
        "total_depenses_mois":  total_depenses_mois,
        "balance":              total_paiements_mois - total_depenses_mois,
        "depenses":             depenses,
    }


# ============================================================
# Export Saisie Grille — EXCEL
# ============================================================
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def saisie_grille_export_excel(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    residence = get_user_residence(request)
    if not residence:
        return Response({"detail": "Aucune résidence."}, status=400)

    try:
        year  = int(request.query_params.get("year",  datetime.date.today().year))
        month = int(request.query_params.get("month", datetime.date.today().month))
    except (TypeError, ValueError):
        return Response({"detail": "Paramètres invalides."}, status=400)

    type_charge = request.query_params.get("type_charge", "CHARGE")
    if type_charge not in ("CHARGE", "FOND"):
        type_charge = "CHARGE"
    appel_id = request.query_params.get("appel_id")
    try:
        appel_id = int(appel_id) if appel_id else None
    except (TypeError, ValueError):
        appel_id = None

    d = _saisie_grille_data(residence, year, month, type_charge, appel_id=appel_id)

    wb = Workbook()
    ws = wb.active
    ws.title = "Saisie Grille"

    thin   = Side(style="thin", color="E2E8F0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    DARK  = "0F172A"
    INDIGO = "4F46E5"
    GREEN  = "059669"
    RED    = "DC2626"
    AMBER  = "D97706"

    def hdr(cell, bg=DARK, color="FFFFFF", size=10, bold=True, center=True):
        cell.fill      = PatternFill("solid", fgColor=bg)
        cell.font      = Font(bold=bold, color=color, size=size)
        cell.alignment = Alignment(horizontal="center" if center else "left", vertical="center", wrap_text=True)
        cell.border    = border

    def dc(ws, row, col, value, fmt=None, bold=False, color=None, center=False):
        c = ws.cell(row=row, column=col, value=value)
        c.border = border
        c.alignment = Alignment(horizontal="center" if center else "left", vertical="center")
        if fmt:   c.number_format = fmt
        if bold:  c.font = Font(bold=True, color=color or "000000")
        elif color: c.font = Font(color=color)
        return c

    tc_label = "Appel de charge" if type_charge == "CHARGE" else "Appel de fond"
    mois_label = MOIS_FR[month - 1]

    # ── Titre ──────────────────────────────────────────────────
    ws.merge_cells("A1:Q1")
    c = ws["A1"]
    c.value = f"SAISIE EN GRILLE — {residence.nom_residence.upper()}"
    c.font  = Font(bold=True, size=14, color=DARK)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22

    ws.merge_cells("A2:Q2")
    ws["A2"].value = f"{tc_label}  ·  {mois_label} {year}"
    ws["A2"].font  = Font(italic=True, color="64748B", size=10)
    ws.append([])  # row 3 blank

    # ── KPI ────────────────────────────────────────────────────
    kpi_row = 4
    for ci, (lbl, val, col) in enumerate([
        ("Paiements reçus", d["total_paiements_mois"], GREEN),
        ("Dépenses",        d["total_depenses_mois"],  RED),
        ("Balance",         d["balance"],              GREEN if d["balance"] >= 0 else RED),
    ], 1):
        col_start = (ci - 1) * 3 + 1
        ws.merge_cells(start_row=kpi_row,   start_column=col_start, end_row=kpi_row,   end_column=col_start + 2)
        ws.merge_cells(start_row=kpi_row+1, start_column=col_start, end_row=kpi_row+1, end_column=col_start + 2)
        lc = ws.cell(kpi_row, col_start, lbl)
        lc.font = Font(bold=True, color="94A3B8", size=9)
        lc.alignment = Alignment(horizontal="center")
        vc = ws.cell(kpi_row+1, col_start, val)
        vc.font = Font(bold=True, color=col, size=13)
        vc.alignment = Alignment(horizontal="center")
        vc.number_format = "#,##0.00"
    ws.append([])  # blank

    # ── En-têtes grille ─────────────────────────────────────────
    hdr_row = kpi_row + 3
    headers = ["Lot", "Propriétaire"] + MOIS_SHORT + ["Total dû", "Total payé", "Reste"]
    col_widths = [8, 26] + [5]*12 + [14, 14, 14]
    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        ws.column_dimensions[ws.cell(hdr_row, ci).column_letter].width = w
        hdr(ws.cell(hdr_row, ci, h), bg=INDIGO)

    # ── Données grille ──────────────────────────────────────────
    total_du_sum = total_paye_sum = total_reste_sum = 0.0
    for ri, lot in enumerate(d["lots"], hdr_row + 1):
        dc(ws, ri, 1, lot["lot"],  bold=True, color=INDIGO, center=True)
        dc(ws, ri, 2, lot["nom"])
        for mi in range(12):
            is_this_month = lot["paid_this_month"][mi]
            is_before     = lot["paid_before"][mi]
            c = ws.cell(ri, 3 + mi, "✓" if (is_this_month or is_before) else "")
            c.border = border
            c.alignment = Alignment(horizontal="center", vertical="center")
            if is_this_month:
                c.font = Font(bold=True, color=GREEN)
                c.fill = PatternFill("solid", fgColor="D1FAE5")
            elif is_before:
                c.font = Font(bold=False, color="6EE7B7")
        dc(ws, ri, 15, lot["total_du"],   fmt="#,##0.00", center=True)
        dc(ws, ri, 16, lot["total_paye"], fmt="#,##0.00", center=True)
        reste = lot["reste"]
        c = ws.cell(ri, 17, reste)
        c.border = border
        c.number_format = "#,##0.00"
        c.alignment = Alignment(horizontal="center")
        c.font = Font(bold=True, color=RED if reste > 0 else GREEN)
        total_du_sum    += lot["total_du"]
        total_paye_sum  += lot["total_paye"]
        total_reste_sum += lot["reste"]

    # Totaux grille
    tr = hdr_row + 1 + len(d["lots"])
    c = ws.cell(tr, 1, "TOTAL")
    c.font = Font(bold=True)
    c.border = border
    ws.merge_cells(start_row=tr, start_column=1, end_row=tr, end_column=14)
    for ci, val in [(15, total_du_sum), (16, total_paye_sum), (17, total_reste_sum)]:
        c = ws.cell(tr, ci, val)
        c.number_format = "#,##0.00"
        c.font = Font(bold=True, color=RED if ci == 17 and total_reste_sum > 0 else GREEN)
        c.border = border
        c.alignment = Alignment(horizontal="center")

    # ── Section dépenses (fusionnée sur 17 colonnes) ────────────
    # Spans : Date=1-2, Libellé=3-9, Catégorie=10-12, Fournisseur=13-15, Montant=16-17
    DEP_SPANS = [(1, 2), (3, 9), (10, 12), (13, 15), (16, 17)]
    DEP_HDR_LABELS = ["Date", "Libellé", "Catégorie", "Fournisseur", "Montant (MAD)"]
    DEP_FIELDS     = ["date", "libelle", "categorie", "fournisseur", "montant"]

    def dep_merge(ws, row, span_idx, value, is_hdr=False, is_num=False, is_bold=False, color=None):
        c1, c2 = DEP_SPANS[span_idx]
        if c2 > c1:
            ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
        cell = ws.cell(row, c1, value)
        cell.border = border
        align_h = "center" if span_idx in (0, 4) else "left"
        cell.alignment = Alignment(horizontal=align_h, vertical="center")
        if is_hdr:
            cell.fill  = PatternFill("solid", fgColor=RED)
            cell.font  = Font(bold=True, color="FFFFFF", size=10)
        elif is_num:
            cell.number_format = "#,##0.00"
            cell.font = Font(bold=is_bold, color=color or "000000")
        elif is_bold:
            cell.font = Font(bold=True, color=color or "000000")
        return cell

    dep_start = tr + 2
    ws.merge_cells(start_row=dep_start, start_column=1, end_row=dep_start, end_column=17)
    c = ws.cell(dep_start, 1, f"DÉPENSES — {mois_label} {year}  ({len(d['depenses'])} entrée(s))")
    c.font = Font(bold=True, size=11, color=RED)
    c.alignment = Alignment(horizontal="left")

    dep_hdr = dep_start + 1
    for si, label in enumerate(DEP_HDR_LABELS):
        dep_merge(ws, dep_hdr, si, label, is_hdr=True)

    for ri, dep in enumerate(d["depenses"], dep_hdr + 1):
        for si, field in enumerate(DEP_FIELDS):
            val = dep[field]
            dep_merge(ws, ri, si, val, is_num=(si == 4))

    # Total dépenses
    tdr = dep_hdr + 1 + len(d["depenses"])
    ws.merge_cells(start_row=tdr, start_column=1, end_row=tdr, end_column=15)
    c = ws.cell(tdr, 1, "TOTAL DÉPENSES")
    c.font = Font(bold=True)
    c.border = border
    c.alignment = Alignment(horizontal="left", vertical="center")
    dep_merge(ws, tdr, 4, d["total_depenses_mois"], is_num=True, is_bold=True, color=RED)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    slug  = re.sub(r"[^a-z0-9]", "_", residence.nom_residence.lower())
    today = datetime.date.today().strftime("%Y%m%d")
    resp  = HttpResponse(buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = f'attachment; filename="grille_{slug}_{year}_{month:02d}_{today}.xlsx"'
    return resp


# ============================================================
# Export Saisie Grille — PDF
# ============================================================
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def saisie_grille_export_pdf(request):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT

    residence = get_user_residence(request)
    if not residence:
        return Response({"detail": "Aucune résidence."}, status=400)

    try:
        year  = int(request.query_params.get("year",  datetime.date.today().year))
        month = int(request.query_params.get("month", datetime.date.today().month))
    except (TypeError, ValueError):
        return Response({"detail": "Paramètres invalides."}, status=400)

    type_charge = request.query_params.get("type_charge", "CHARGE")
    if type_charge not in ("CHARGE", "FOND"):
        type_charge = "CHARGE"
    appel_id_pdf = request.query_params.get("appel_id")
    try:
        appel_id_pdf = int(appel_id_pdf) if appel_id_pdf else None
    except (TypeError, ValueError):
        appel_id_pdf = None

    d = _saisie_grille_data(residence, year, month, type_charge, appel_id=appel_id_pdf)

    def money(v): return f"{float(v):,.2f}"

    DARK   = colors.HexColor("#0F172A")
    INDIGO = colors.HexColor("#4F46E5")
    GREEN  = colors.HexColor("#059669")
    RED    = colors.HexColor("#DC2626")
    AMBER  = colors.HexColor("#D97706")
    LGRAY  = colors.HexColor("#F8FAFC")
    BORDER = colors.HexColor("#E2E8F0")

    buf  = io.BytesIO()
    doc  = SimpleDocTemplate(buf, pagesize=landscape(A4),
                             leftMargin=1.2*cm, rightMargin=1.2*cm,
                             topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    story  = []

    def para(text, size=10, bold=False, color=DARK, align="LEFT"):
        alignment = {"LEFT": 0, "CENTER": 1, "RIGHT": 2}.get(align, 0)
        style = ParagraphStyle("x", fontSize=size, fontName="Helvetica-Bold" if bold else "Helvetica",
                               textColor=color, alignment=alignment, spaceAfter=2)
        return Paragraph(text, style)

    tc_label  = "Appel de charge" if type_charge == "CHARGE" else "Appel de fond"
    mois_label = MOIS_FR[month - 1]

    # ── Titre ──────────────────────────────────────────────────
    story.append(para(f"SAISIE EN GRILLE — {residence.nom_residence.upper()}", size=14, bold=True))
    story.append(para(f"{tc_label}  ·  {mois_label} {year}", size=10, color=colors.HexColor("#64748B")))
    story.append(Spacer(1, 0.4*cm))

    # ── KPI ────────────────────────────────────────────────────
    kpi_data = [
        ["Paiements reçus", "Dépenses", "Balance"],
        [money(d["total_paiements_mois"]) + " MAD",
         money(d["total_depenses_mois"])  + " MAD",
         money(d["balance"])              + " MAD"],
    ]
    bal_color = GREEN if d["balance"] >= 0 else RED
    kpi_table = Table(kpi_data, colWidths=[8*cm, 8*cm, 8*cm])
    kpi_table.setStyle(TableStyle([
        ("BACKGROUND",   (0,0), (-1,0), LGRAY),
        ("FONTNAME",     (0,0), (-1,0), "Helvetica"),
        ("FONTSIZE",     (0,0), (-1,0), 8),
        ("TEXTCOLOR",    (0,0), (-1,0), colors.HexColor("#94A3B8")),
        ("FONTNAME",     (0,1), (-1,1), "Helvetica-Bold"),
        ("FONTSIZE",     (0,1), (-1,1), 12),
        ("TEXTCOLOR",    (0,1), (0,1), GREEN),
        ("TEXTCOLOR",    (1,1), (1,1), RED),
        ("TEXTCOLOR",    (2,1), (2,1), bal_color),
        ("ALIGN",        (0,0), (-1,-1), "CENTER"),
        ("VALIGN",       (0,0), (-1,-1), "MIDDLE"),
        ("ROWBACKGROUNDS",(0,0),(-1,-1), [LGRAY, colors.white]),
        ("BOX",          (0,0), (-1,-1), 0.5, BORDER),
        ("INNERGRID",    (0,0), (-1,-1), 0.5, BORDER),
        ("TOPPADDING",   (0,0), (-1,-1), 6),
        ("BOTTOMPADDING",(0,0), (-1,-1), 6),
    ]))
    story.append(kpi_table)
    story.append(Spacer(1, 0.5*cm))

    # ── Grille paiements ────────────────────────────────────────
    story.append(para("GRILLE DE PAIEMENT", size=10, bold=True, color=INDIGO))
    story.append(Spacer(1, 0.2*cm))

    grille_headers = ["Lot", "Propriétaire"] + MOIS_SHORT + ["Total dû", "Payé", "Reste"]
    grille_rows = [grille_headers]
    for lot in d["lots"]:
        row = [lot["lot"], lot["nom"]]
        for mi in range(12):
            row.append("✓" if (lot["paid_this_month"][mi] or lot["paid_before"][mi]) else "")
        row += [money(lot["total_du"]), money(lot["total_paye"]), money(lot["reste"])]
        grille_rows.append(row)

    # Total row
    grille_rows.append(
        ["TOTAL", ""] + [""] * 12 +
        [money(sum(l["total_du"]   for l in d["lots"])),
         money(sum(l["total_paye"] for l in d["lots"])),
         money(sum(l["reste"]      for l in d["lots"]))]
    )

    pw = 27.7 * cm  # landscape A4 usable width
    col_w = [1.2*cm, 5.5*cm] + [0.7*cm]*12 + [2.3*cm, 2.3*cm, 2.3*cm]
    grille_table = Table(grille_rows, colWidths=col_w, repeatRows=1)

    ts = TableStyle([
        ("BACKGROUND",   (0,0), (-1,0),  INDIGO),
        ("TEXTCOLOR",    (0,0), (-1,0),  colors.white),
        ("FONTNAME",     (0,0), (-1,0),  "Helvetica-Bold"),
        ("FONTSIZE",     (0,0), (-1,-1), 7),
        ("ALIGN",        (0,0), (-1,-1), "CENTER"),
        ("ALIGN",        (1,1), (1,-1),  "LEFT"),
        ("VALIGN",       (0,0), (-1,-1), "MIDDLE"),
        ("ROWBACKGROUNDS",(0,1),(-1,-2), [colors.white, LGRAY]),
        ("BOX",          (0,0), (-1,-1), 0.5, BORDER),
        ("INNERGRID",    (0,0), (-1,-1), 0.3, BORDER),
        ("TOPPADDING",   (0,0), (-1,-1), 3),
        ("BOTTOMPADDING",(0,0), (-1,-1), 3),
        # Total row
        ("BACKGROUND",   (0,-1), (-1,-1), LGRAY),
        ("FONTNAME",     (0,-1), (-1,-1), "Helvetica-Bold"),
    ])
    # Color ✓ per row: bright green = paid this month, muted = paid before
    MUTED_GREEN = colors.HexColor("#10B981")
    CELL_BG     = colors.HexColor("#D1FAE5")
    for ri, lot in enumerate(d["lots"], 1):
        for mi in range(12):
            if lot["paid_this_month"][mi]:
                ts.add("TEXTCOLOR",  (2+mi, ri), (2+mi, ri), GREEN)
                ts.add("BACKGROUND", (2+mi, ri), (2+mi, ri), CELL_BG)
            elif lot["paid_before"][mi]:
                ts.add("TEXTCOLOR",  (2+mi, ri), (2+mi, ri), MUTED_GREEN)
        reste_col = -1
        ts.add("TEXTCOLOR", (reste_col, ri), (reste_col, ri),
               RED if lot["reste"] > 0 else GREEN)
    grille_table.setStyle(ts)
    story.append(grille_table)
    story.append(Spacer(1, 0.6*cm))

    # ── Dépenses ────────────────────────────────────────────────
    story.append(para(f"DÉPENSES — {mois_label} {year}  ({len(d['depenses'])} entrée(s))", size=10, bold=True, color=RED))
    story.append(Spacer(1, 0.2*cm))

    dep_headers = ["Date", "Libellé", "Catégorie", "Fournisseur", "Montant (MAD)"]
    dep_rows = [dep_headers]
    for dep in d["depenses"]:
        dep_rows.append([dep["date"], dep["libelle"], dep["categorie"], dep["fournisseur"], money(dep["montant"])])
    dep_rows.append(["", "TOTAL", "", "", money(d["total_depenses_mois"])])

    dep_table = Table(dep_rows, colWidths=[2.2*cm, 8*cm, 4*cm, 5*cm, 3*cm], repeatRows=1)
    dep_table.setStyle(TableStyle([
        ("BACKGROUND",   (0,0), (-1,0), RED),
        ("TEXTCOLOR",    (0,0), (-1,0), colors.white),
        ("FONTNAME",     (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE",     (0,0), (-1,-1), 8),
        ("ALIGN",        (0,0), (-1,-1), "LEFT"),
        ("ALIGN",        (4,0), (4,-1),  "RIGHT"),
        ("VALIGN",       (0,0), (-1,-1), "MIDDLE"),
        ("ROWBACKGROUNDS",(0,1),(-1,-2), [colors.white, LGRAY]),
        ("BOX",          (0,0), (-1,-1), 0.5, BORDER),
        ("INNERGRID",    (0,0), (-1,-1), 0.3, BORDER),
        ("TOPPADDING",   (0,0), (-1,-1), 3),
        ("BOTTOMPADDING",(0,0), (-1,-1), 3),
        ("FONTNAME",     (0,-1), (-1,-1), "Helvetica-Bold"),
        ("TEXTCOLOR",    (4,-1), (4,-1),  RED),
    ]))
    story.append(dep_table)

    doc.build(story)
    buf.seek(0)
    slug  = re.sub(r"[^a-z0-9]", "_", residence.nom_residence.lower())
    today = datetime.date.today().strftime("%Y%m%d")
    resp  = HttpResponse(buf.read(), content_type="application/pdf")
    resp["Content-Disposition"] = f'attachment; filename="grille_{slug}_{year}_{month:02d}_{today}.pdf"'
    return resp
